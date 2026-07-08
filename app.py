import streamlit as st
import xml.etree.ElementTree as ET
from datetime import datetime
import re
import zipfile
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_DISPONIVEL = True
except ImportError:
    EXCEL_DISPONIVEL = False

VERSAO = "V5.1-FINAL"
DATA_CADASTRO_FIXO = "01/01/2020"

def apply_tr_theme():
    st.markdown("""
        <style>
        html, body, [class*="css"] { font-family: 'Segoe UI', 'Arial', sans-serif; color: #444444; }
        h1, h2, h3 { color: #FF8000; font-weight: 700; }
        section[data-testid="stSidebar"] { background-color: #444444; color: #FFFFFF; }
        section[data-testid="stSidebar"] * { color: #FFFFFF !important; }
        .stButton > button { background-color: #FF8000; color: #FFFFFF; border: none; border-radius: 4px; font-weight: bold; }
        .stButton > button:hover { background-color: #D64001; }
        .stDownloadButton > button { background-color: #FF8000; color: #FFFFFF; border: none; border-radius: 4px; font-weight: bold; }
        .stDownloadButton > button:hover { background-color: #D64001; }
        hr { border-color: #FF8000; }
        [data-testid="metric-container"] { background-color: #E9E9E9; border-left: 4px solid #FF8000; border-radius: 4px; padding: 10px; }
        .instrucoes-box { background-color: #E9E9E9; border-left: 4px solid #FF8000; border-radius: 4px; padding: 16px 20px; margin: 12px 0; color: #444444; }
        .instrucoes-box h4 { color: #FF8000; margin-top: 14px; margin-bottom: 6px; }
        .instrucoes-box h4:first-child { margin-top: 0; }
        .cnpj-badge { background-color: #444444; border: 1px solid #FF8000; border-radius: 4px; padding: 6px 12px; font-family: Consolas, monospace; font-size: 13px; display: inline-block; margin: 4px 0; }
        .info-origem { background-color: #FFF3E0; border-left: 3px solid #FF8000; border-radius: 3px; padding: 8px 12px; font-size: 12px; color: #444444; margin: 6px 0; }
        .zip-info { background-color: #E3F2FD; border-left: 4px solid #1565C0; border-radius: 4px; padding: 10px 14px; margin: 8px 0; font-size: 13px; color: #1565C0; }
        .zip-warn { background-color: #FFF8E1; border-left: 4px solid #F9A825; border-radius: 4px; padding: 10px 14px; margin: 8px 0; font-size: 13px; color: #795548; }
        </style>
    """, unsafe_allow_html=True)

st.set_page_config(
    page_title="Dominio Sistemas | Thomson Reuters",
    page_icon="🟠",
    layout="wide",
    initial_sidebar_state="expanded"
)
apply_tr_theme()

st.markdown(f"""
    <div style="background:#444444; padding:24px 28px 18px 28px; border-radius:8px;
                border-top:6px solid #FF8000; margin-bottom:28px;">
        <h2 style="color:#FF8000; margin:0; font-family:'Segoe UI',Arial,sans-serif;">
            🔄 Conversor XML NF-e → Dominio Sistemas &nbsp;|&nbsp; {VERSAO}
        </h2>
        <p style="color:#DDDDDD; margin:6px 0 0 0; font-family:'Segoe UI',Arial,sans-serif;">
            Converte XML de NF-e para leiaute padrao de importacao do
            <strong>Dominio Sistemas</strong>.
            Saida em <strong>ANSI (Latin-1)</strong>. Aceita <strong>XML</strong>
            ou <strong>ZIP</strong>.
        </p>
    </div>
""", unsafe_allow_html=True)

NS = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

TABELA_GRUPOS = {
    0:  "Automatico (por CFOP/NCM)",
    1:  "GERAL",
    2:  "MERCADORIA PARA REVENDA",
    3:  "MATERIA PRIMA",
    4:  "EMBALAGENS",
    5:  "PRODUTO EM PROCESSO",
    6:  "PRODUTO ACABADO",
    7:  "SUBPRODUTO",
    8:  "PRODUTOS INTERMEDIARIOS",
    9:  "MATERIAL DE USO E CONSUMO",
    10: "ATIVO IMOBILIZADO",
    11: "SERVICOS",
    12: "OUTROS INSUMOS",
    100:"PRODUTOS NOVOS - ENTRADAS",
    500:"PRODUTOS NOVOS - SAIDAS",
}

# ─────────────────────────────────────────────────────────────────────────────
# MUDANÇA 1 — Tabelas de CST PIS/COFINS e função obter_cst_entrada
# Substituem completamente cst_pis_efetivo / cst_cof_efetivo
# ─────────────────────────────────────────────────────────────────────────────

# Tabela de CST PIS/COFINS para ENTRADAS
# Chave = CST normalizado (2 dígitos, string)
# Valor = descrição resumida para referência
TABELA_CST_PIS_COFINS_ENTRADA = {
    "01": "01 - Op. Tributavel - Aliq. Normal (cumulativo/nao-cumulativo)",
    "02": "02 - Op. Tributavel - Aliq. Diferenciada",
    "03": "03 - Op. Tributavel - Qtde (pauta)",
    "04": "04 - Op. Tributavel - Monofasica - Aliq. Zero",
    "05": "05 - Op. Tributavel - Substituicao Tributaria",
    "06": "06 - Op. Tributavel - Aliq. Zero",
    "07": "07 - Op. Isenta de Contribuicao",
    "08": "08 - Op. Sem Incidencia da Contribuicao",
    "09": "09 - Op. com Suspensao da Contribuicao",
    "49": "49 - Outras Operacoes de Entrada",
    "50": "50 - Op. com Direito a Credito - Vinc. Excl. a Rec. Nao-Cumulativo",
    "51": "51 - Op. com Direito a Credito - Vinc. Excl. a Rec. Cumulativo",
    "52": "52 - Op. com Direito a Credito - Vinc. a Rec. Nao-Cumulativo e Cumulativo",
    "53": "53 - Op. com Direito a Credito - Vinc. a Rec. Tributadas e Nao-Tributadas",
    "54": "54 - Op. com Direito a Credito - Vinc. a Rec. Tributadas e de Exportacao",
    "55": "55 - Op. com Direito a Credito - Vinc. a Rec. Nao-Tributadas e Exportacao",
    "56": "56 - Op. com Direito a Credito - Vinc. a Rec. Tributadas, Nao-Tributadas e Exportacao",
    "60": "60 - Credito Presumido - Op. de Aquisicao Vinc. Excl. a Rec. Tributada Nao-Cumulativo",
    "61": "61 - Credito Presumido - Op. de Aquisicao Vinc. Excl. a Rec. Nao-Tributada",
    "62": "62 - Credito Presumido - Op. de Aquisicao Vinc. Excl. a Rec. de Exportacao",
    "63": "63 - Credito Presumido - Op. de Aquisicao Vinc. a Rec. Nao-Tributadas e Exportacao",
    "64": "64 - Credito Presumido - Op. de Aquisicao Vinc. a Rec. Tributadas e Nao-Tributadas",
    "65": "65 - Credito Presumido - Op. de Aquisicao Vinc. a Rec. Tributadas e de Exportacao",
    "66": "66 - Credito Presumido - Op. de Aquisicao Vinc. a Rec. Trib., Nao-Trib. e Exportacao",
    "67": "67 - Credito Presumido - Outras Operacoes",
    "70": "70 - Op. de Aquisicao sem Direito a Credito",
    "71": "71 - Op. de Aquisicao com Isencao",
    "72": "72 - Op. de Aquisicao com Suspensao",
    "73": "73 - Op. de Aquisicao a Aliq. Zero",
    "74": "74 - Op. de Aquisicao sem Incidencia da Contribuicao",
    "75": "75 - Op. de Aquisicao por Substituicao Tributaria",
    "98": "98 - Outras Operacoes de Entrada",
    "99": "99 - Outras Operacoes",
}

# Tabela de CST PIS/COFINS para SAÍDAS
TABELA_CST_PIS_COFINS_SAIDA = {
    "01": "01 - Op. Tributavel - Aliq. Normal",
    "02": "02 - Op. Tributavel - Aliq. Diferenciada",
    "03": "03 - Op. Tributavel - Qtde (pauta)",
    "04": "04 - Op. Tributavel - Monofasica - Aliq. Zero",
    "05": "05 - Op. Tributavel - Substituicao Tributaria",
    "06": "06 - Op. Tributavel - Aliq. Zero",
    "07": "07 - Op. Isenta de Contribuicao",
    "08": "08 - Op. Sem Incidencia da Contribuicao",
    "09": "09 - Op. com Suspensao da Contribuicao",
    "49": "49 - Outras Operacoes de Saida",
    "99": "99 - Outras Operacoes",
}

# DE/PARA: CST de saída → CST de entrada equivalente (para 0110 campo cst_s)
CST_ENTRADA_SAIDA = {
    "01": "50",
    "02": "50",
    "03": "50",
    "04": "70",
    "05": "75",
    "06": "73",
    "07": "71",
    "08": "74",
    "09": "72",
    "49": "98",
    "50": "01",
    "51": "02",
    "52": "08",
    "73": "06",
    "74": "08",
    "70": "04",
    "99": "49",
}


def obter_cst_entrada(cst_norm: str, tem_direito_credito: bool) -> str:
    """
    Recebe o CST já normalizado (zfill(2)) lido do XML.

    Regra:
    - Se tem_direito_credito=True e o CST do XML for um CST de SAÍDA (01-09, 49, 99),
      converte para o CST de ENTRADA com direito a crédito correspondente
      usando CST_ENTRADA_SAIDA.
    - Se tem_direito_credito=False, devolve o CST do XML sem transformação.
    - Se o CST já for um CST de ENTRADA (50-98), devolve sem transformação
      independentemente do flag.
    """
    # CSTs que são tipicamente de saída e precisam ser convertidos para entrada
    CST_SAIDA_SET = {"01","02","03","04","05","06","07","08","09","49","99"}

    if tem_direito_credito and cst_norm in CST_SAIDA_SET:
        return CST_ENTRADA_SAIDA.get(cst_norm, cst_norm)

    # Sem direito a crédito: devolve o CST do XML sem transformação
    return cst_norm


# Mantidas como wrappers para compatibilidade com chamadas existentes no Excel
def cst_pis_efetivo(cst_xml: str, aliq_item: float, aliq_padrao: float) -> str:
    """Legado — usado apenas em extrair_dados_impostos_itens (Excel)."""
    cst_norm = str(cst_xml).strip().zfill(2)
    if aliq_padrao > 0 and aliq_item > 0 and aliq_item < aliq_padrao:
        return "73"
    return cst_norm


def cst_cof_efetivo(cst_xml: str, aliq_item: float, aliq_padrao: float) -> str:
    """Legado — usado apenas em extrair_dados_impostos_itens (Excel)."""
    cst_norm = str(cst_xml).strip().zfill(2)
    if aliq_padrao > 0 and aliq_item > 0 and aliq_item < aliq_padrao:
        return "73"
    return cst_norm


# ─────────────────────────────────────────────────────────────────────────────
# DE/PARA CFOP → ACUMULADOR
# ─────────────────────────────────────────────────────────────────────────────
CFOP_ACUMULADOR_PADRAO = {
    "3101": "1108",
    "3102": "1157",
    "3126": "",
    "3127": "",
    "3201": "",
    "3202": "",
    "3211": "",
    "3251": "",
    "3301": "",
    "3351": "",
    "3352": "",
    "3503": "",
    "3551": "",
    "3553": "",
    "3556": "",
    "3930": "",
    "3949": "",
}

CFOP_DESCRICAO = {
    "3101": "Compra p/ industrializacao - importacao direta",
    "3102": "Compra p/ comercializacao - importacao direta",
    "3126": "Compra p/ uso na prestacao de servico",
    "3127": "Compra p/ uso na prestacao de servico (via interior)",
    "3201": "Devolucao de venda p/ industrializacao",
    "3202": "Devolucao de venda p/ comercializacao",
    "3211": "Devolucao de venda de energia eletrica",
    "3251": "Compra de energia eletrica p/ distribuicao",
    "3301": "Aquisicao de servico de comunicacao",
    "3351": "Compra p/ industrializacao sob encomenda",
    "3352": "Compra p/ comercializacao sob encomenda",
    "3503": "Devolucao de venda de mercadoria exportada",
    "3551": "Compra de bem p/ ativo imobilizado",
    "3553": "Compra de material p/ uso e consumo",
    "3556": "Compra p/ industrializacao originada de encomenda",
    "3930": "Lancamento sob regime aduaneiro especial",
    "3949": "Outra entrada nao especificada",
}

PAISES_BACEN_PARA_DOMINIO = {
    "0132":1,"7560":2,"0175":3,"0230":4,"0370":5,"0400":6,"0418":7,"0434":8,"0477":9,"0531":10,
    "0590":11,"0639":12,"0647":13,"0655":14,"0698":15,"0728":16,"0736":17,"0779":18,"0809":19,"0817":20,
    "0833":21,"0850":22,"0876":23,"0884":24,"0906":25,"0930":26,"0973":27,"0981":28,"1015":29,"1058":30,
    "1082":31,"1112":32,"1155":33,"1198":34,"1279":35,"1376":36,"1414":37,"1457":38,"1490":39,"1504":40,
    "1511":41,"1546":42,"1554":43,"1589":44,"1600":45,"1635":46,"1651":47,"1694":48,"1716":49,"1750":50,
    "1767":51,"1792":52,"1830":53,"1872":54,"1880":55,"1902":56,"1937":57,"1953":58,"1961":59,"1988":60,
    "2003":61,"2070":62,"2100":63,"2127":64,"2151":65,"2186":66,"2291":67,"2321":68,"2356":69,"2399":70,
    "2402":71,"2445":72,"2453":73,"2461":74,"2484":75,"2496":76,"2518":77,"2534":78,"2550":79,"2593":80,
    "2674":81,"2712":82,"2755":83,"2810":84,"2836":85,"2895":86,"2917":87,"2933":88,"2976":89,"3018":90,
    "3050":91,"3093":92,"3107":93,"3131":94,"3174":95,"3204":96,"3212":97,"3255":98,"3298":99,"3310":100,
    "3344":101,"3352":102,"3360":103,"3417":104,"3450":105,"3484":106,"3492":107,"3557":108,"3573":109,"3611":110,
    "3654":111,"3697":112,"3727":113,"3751":114,"3794":115,"3808":116,"3832":117,"3867":118,"3883":119,"3913":120,
    "3964":121,"3999":122,"4030":123,"4111":124,"4200":125,"4235":126,"4260":127,"4278":128,"4316":129,"4340":130,
    "4383":131,"4405":132,"4421":133,"4456":134,"4472":135,"4499":136,"4502":137,"4553":138,"4588":139,"4618":140,
    "4642":141,"4677":142,"4723":143,"4740":144,"4766":145,"4774":146,"4855":147,"4880":148,"4936":149,"4944":150,
    "4952":151,"5010":152,"5053":153,"5070":154,"5088":155,"5096":156,"5177":157,"5215":158,"5258":159,"5282":160,
    "5380":161,"5428":162,"5487":163,"5568":164,"5665":165,"5738":166,"5754":167,"5762":168,"5800":169,"5835":170,
    "5851":171,"5894":172,"5932":173,"5991":174,"6033":175,"6076":176,"6114":177,"6238":178,"6254":179,"6289":180,
    "6300":181,"6327":182,"6408":183,"6432":184,"6459":185,"6505":186,"6513":187,"6548":188,"6580":189,"6599":190,
    "6645":191,"6700":192,"6750":193,"6769":194,"6777":195,"6781":196,"6793":197,"6807":198,"6815":199,"6858":200,
    "6866":201,"6904":202,"6912":203,"7005":204,"7030":205,"7056":206,"7102":207,"7153":208,"7200":209,"7285":210,
    "7315":211,"7358":212,"7370":213,"7412":214,"7447":215,"7455":216,"7501":217,"7544":218,"7552":219,"7590":220,
    "7595":221,"7641":222,"7676":223,"7706":224,"7722":225,"7757":226,"7765":227,"7773":228,"7781":229,"7820":230,
    "7838":231,"7889":232,"7919":233,"7951":234,"7994":235,"8001":236,"8052":237,"8079":238,"8087":239,"8109":240,
    "8150":241,"8168":242,"8176":243,"8230":244,"8273":245,"8281":248,"8311":249,"8338":250,"8346":251,"8451":252,
    "8478":253,"8486":254,"8494":255,"8508":256,"8516":257,"8524":258,"8532":259,"8540":260,"8559":261,"8567":262,
    "8575":263,"8583":264,"8591":265,"8605":266,"8613":267,"8630":268,
}

PAISES_NOME_PARA_DOMINIO = {
    "AFEGANISTAO":1,"AFRICA DO SUL":2,"ALBANIA":3,"ALEMANHA":4,"ANDORRA":5,"ANGOLA":6,
    "ANGUILLA":7,"ANTIGUA E BARBUDA":8,"ANTILHAS HOLANDESAS":9,"ARABIA SAUDITA":10,
    "ARGELIA":11,"ARGENTINA":12,"ARMENIA":13,"ARUBA":14,"AUSTRALIA":15,"AUSTRIA":16,
    "AZERBAIJAO":17,"BAHAMAS":18,"BAHREIN":19,"BANGLADESH":20,"BARBADOS":21,"BELARUS":22,
    "BELGICA":23,"BELIZE":24,"BENIN":25,"BERMUDAS":26,"BOLIVIA":27,"BOSNIA":28,
    "BOTSUANA":29,"BRASIL":30,"BRUNEI":31,"BULGARIA":32,"BURKINA FASO":33,"BURUNDI":34,
    "BUTAO":35,"CABO VERDE":36,"CAMAROES":37,"CAMBOJA":38,"CANADA":39,"CANARIAS":41,
    "CATAR":42,"CAYMAN":43,"CAZAQUISTAO":44,"CHADE":45,"CHILE":46,"CHINA":47,"CHIPRE":48,
    "CHRISTMAS":49,"CINGAPURA":50,"SINGAPURA":50,"COCOS":51,"COLOMBIA":52,"COMORES":53,
    "CONGO":54,"COOK":56,"COREIA DO NORTE":57,"COREIA DO SUL":58,"COSTA DO MARFIM":59,
    "COSTA RICA":60,"KUWAIT":61,"CROACIA":62,"CUBA":63,"DINAMARCA":64,"DJIBUTI":65,
    "DOMINICA":66,"EGITO":67,"EL SALVADOR":68,"EMIRADOS ARABES UNIDOS":69,"EQUADOR":70,
    "ERITREIA":71,"ESCOCIA":72,"ESLOVACA":73,"ESLOVENIA":74,"ESPANHA":75,
    "ESTADOS UNIDOS":76,"ESTONIA":77,"ETIOPIA":78,"FALKLAND":79,"FEROE":80,"FIJI":81,
    "FILIPINAS":82,"FINLANDIA":83,"FORMOSA":84,"TAIWAN":84,"FRANCA":85,"GABAO":86,
    "GALES":87,"GAMBIA":88,"GANA":89,"GEORGIA":90,"GIBRALTAR":91,"GRA-BRETANHA":92,
    "GRANADA":93,"GRECIA":94,"GROENLANDIA":95,"GUADALUPE":96,"GUAM":97,"GUATEMALA":98,
    "GUIANA":99,"GUIANA FRANCESA":100,"GUINE":101,"GUINE-BISSAU":102,"GUINE-EQUATORIAL":103,
    "HAITI":104,"HOLANDA":105,"PAISES BAIXOS":105,"HONDURAS":106,"HONG KONG":107,
    "HUNGRIA":108,"IEMEN":109,"INDIA":110,"INDONESIA":111,"INGLATERRA":112,"IRA":113,
    "IRAQUE":114,"IRLANDA":115,"IRLANDA DO NORTE":116,"ISLANDIA":117,"ISRAEL":118,
    "ITALIA":119,"SERVIA":120,"JAMAICA":121,"JAPAO":122,"JOHNSTON":123,"JORDANIA":124,
    "KIRIBATI":125,"LAOS":126,"LEBUAN":127,"LESOTO":128,"LETONIA":129,"LIBANO":130,
    "LIBERIA":131,"LIBIA":132,"LIECHTENSTEIN":133,"LITUANIA":134,"LUXEMBURGO":135,
    "MACAU":136,"MACEDONIA DO NORTE":137,"MADAGASCAR":138,"MADEIRA":139,"MALASIA":140,
    "MALAVI":141,"MALDIVAS":142,"MALI":143,"MALTA":144,"MAN":145,"MARIANAS DO NORTE":146,
    "MARROCOS":147,"MARSHALL":148,"MARTINICA":149,"MAURICIO":150,"MAURITANIA":151,
    "MEXICO":152,"MIANMAR":153,"BIRMANIA":153,"MICRONESIA":154,"MIDWAY":155,
    "MOCAMBIQUE":156,"MOLDAVIA":157,"MONACO":158,"MONGOLIA":159,"MONTSERRAT":160,
    "NAMIBIA":161,"NAURU":162,"NEPAL":163,"NICARAGUA":164,"NIGER":165,"NIGERIA":166,
    "NIUE":167,"NORFOLK":168,"NORUEGA":169,"NOVA CALEDONIA":170,"NOVA ZELANDIA":171,
    "OMA":172,"PALAU":173,"PANAMA":174,"PAPUA NOVA GUINE":175,"PAQUISTAO":176,
    "PARAGUAI":177,"PERU":178,"PITCAIRN":179,"POLINESIA FRANCESA":180,"POLONIA":181,
    "PORTO RICO":182,"PORTUGAL":183,"QUENIA":184,"QUIRGUIZ":185,"REINO UNIDO":186,
    "REPUBLICA CENTRO-AFRICANA":187,"REPUBLICA DOMINICANA":188,"REUNIAO":189,
    "ROMENIA":190,"RUANDA":191,"RUSSIA":192,"SAARA OCIDENTAL":193,"SALOMAO":194,
    "SAMOA":195,"SAMOA AMERICANA":196,"SAN MARINO":197,"SANTA HELENA":198,
    "SANTA LUCIA":199,"SAO CRISTOVAO E NEVES":200,"SAO PEDRO E MIQUELON":201,
    "SAO TOME E PRINCIPE":202,"SAO VICENTE E GRANADINA":203,"SENEGAL":204,
    "SERRA LEOA":205,"SEYCHELLE":206,"SIRIA":207,"SOMALIA":208,"SRI LANKA":209,
    "ESWATINI":210,"SUAZILANDIA":210,"SUDAO":211,"SUECIA":212,"SUICA":213,
    "SURINAME":214,"TADJIQUISTAO":215,"TAILANDIA":216,"TANZANIA":217,"TCHECA":218,
    "TERRITORIO BRITANICO":219,"TIMOR LESTE":220,"TOGO":221,"TONGA":222,"TOQUELAU":223,
    "TRINIDAD E TOBAGO":224,"TUNISIA":225,"TURCAS E CAICOS":226,"TURCOMENISTAO":227,
    "TURQUIA":228,"TUVALU":229,"UCRANIA":230,"UGANDA":231,"URUGUAI":232,
    "UZBEQUISTAO":233,"VANUATU":234,"VATICANO":235,"VENEZUELA":236,"VIETNA":237,
    "VIRGENS BRITANICAS":238,"VIRGENS EUA":239,"WAKE":240,"WALLIS E FUTUNA":241,
    "ZAMBIA":242,"ZIMBABUE":243,"ZONA DO CANAL DO PANAMA":244,"MONTENEGRO":245,
    "QATAR":249,"SAINT KITTS E NEVIS":250,"CURACAO":256,"MAYOTTE":261,
    "PALESTINA":266,"SUDAO DO SUL":267,
}

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def sanitizar_xml_bytes(raw: bytes) -> bytes:
    if not raw:
        return raw
    if raw.startswith(b'\xef\xbb\xbf'):
        raw = raw[3:]
    elif raw.startswith(b'\xff\xfe') or raw.startswith(b'\xfe\xff'):
        raw = raw[2:]
    return raw

def resolver_codigo_pais_dominio(c_pais_xml: str, x_pais_xml: str) -> str:
    c_norm = (c_pais_xml or "").strip().zfill(4)
    if c_norm in PAISES_BACEN_PARA_DOMINIO:
        return str(PAISES_BACEN_PARA_DOMINIO[c_norm])
    if x_pais_xml:
        nome = x_pais_xml.upper().strip()
        if nome in PAISES_NOME_PARA_DOMINIO:
            return str(PAISES_NOME_PARA_DOMINIO[nome])
        for chave, cod in PAISES_NOME_PARA_DOMINIO.items():
            if chave in nome or nome in chave:
                return str(cod)
    return c_pais_xml or ""

def get_text(element, path: str, default: str = "") -> str:
    if element is None:
        return default
    found = element.find(path, NS)
    if found is not None and found.text:
        return found.text.strip()
    return default

def fmt_decimal(value: str, decimals: int = 2) -> str:
    if not value:
        return ""
    try:
        v = float(str(value).replace(",", "."))
        return f"{v:.{decimals}f}".replace(".", ",")
    except (ValueError, TypeError):
        return str(value)

def fmt_date(iso_date: str) -> str:
    if not iso_date:
        return ""
    try:
        return datetime.strptime(iso_date[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return iso_date

def pipe_join(fields: list) -> str:
    return "|" + "|".join(str(f) for f in fields) + "|"

def encode_ansi(conteudo: str) -> bytes:
    resultado = []
    for char in conteudo:
        try:
            resultado.append(char.encode("latin-1"))
        except UnicodeEncodeError:
            resultado.append(b"?")
    return b"".join(resultado)

def somente_numeros(valor: str) -> str:
    return re.sub(r"[^0-9]", "", valor or "")

def extrair_chave_nfe(nfe_root) -> str:
    inf_nfe = nfe_root.find("nfe:infNFe", NS)
    if inf_nfe is None:
        return ""
    id_attr = inf_nfe.get("Id", "")
    chave   = re.sub(r"^NFe", "", id_attr)
    chave_n = re.sub(r"[^0-9]", "", chave)
    return chave_n if len(chave_n) == 44 else chave

def safe_float(v: str) -> float:
    try:
        return float(str(v).replace(",", ".") or "0")
    except (ValueError, TypeError):
        return 0.0

def parse_xml_seguro(raw_bytes: bytes):
    if not raw_bytes:
        return None, "Arquivo vazio (0 bytes)"
    limpo = sanitizar_xml_bytes(raw_bytes)
    try:
        return ET.fromstring(limpo), None
    except ET.ParseError:
        pass
    try:
        texto = limpo.decode("utf-8", errors="replace")
        return ET.fromstring(texto.encode("utf-8")), None
    except ET.ParseError:
        pass
    try:
        texto = limpo.decode("latin-1", errors="replace")
        return ET.fromstring(texto.encode("utf-8")), None
    except ET.ParseError as e:
        return None, f"XML invalido: {e}"

# ─────────────────────────────────────────────
# DE/PARA CFOP → ACUMULADOR
# ─────────────────────────────────────────────
def extrair_cfops_importacao_dos_arquivos(arquivos: list) -> list:
    cfops = set()
    for arq in arquivos:
        root, erro = parse_xml_seguro(arq["bytes"])
        if root is None:
            continue
        nfe = root.find("nfe:NFe", NS) or root
        for det in nfe.findall("nfe:infNFe/nfe:det", NS):
            cfop = get_text(det.find("nfe:prod", NS), "nfe:CFOP").strip()
            if cfop.startswith("3"):
                cfops.add(cfop)
    return sorted(cfops)

def resolver_acumulador_por_cfop(cfop: str,
                                  mapa_customizado: dict,
                                  acumulador_fallback: str = "1157") -> str:
    if cfop in mapa_customizado:
        return mapa_customizado[cfop]
    if cfop in CFOP_ACUMULADOR_PADRAO:
        return CFOP_ACUMULADOR_PADRAO[cfop]
    return acumulador_fallback

# ─────────────────────────────────────────────
# PIS/COFINS — alíquota padrão da nota
# ─────────────────────────────────────────────
def calcular_aliquotas_padrao_nota(nfe_root) -> tuple:
    det_list = nfe_root.findall("nfe:infNFe/nfe:det", NS)
    pis_vals, cof_vals = [], []
    for det in det_list:
        imp = det.find("nfe:imposto", NS)
        if imp is None:
            continue
        pis_node = imp.find("nfe:PIS", NS)
        if pis_node is not None:
            for pt in ["PISAliq","PISQtde","PISNT","PISOutr"]:
                pn = pis_node.find(f"nfe:{pt}", NS)
                if pn is not None:
                    v = safe_float(get_text(pn,"nfe:pPIS") or get_text(pn,"nfe:vAliqProd"))
                    if v > 0:
                        pis_vals.append(v)
                    break
        cof_node = imp.find("nfe:COFINS", NS)
        if cof_node is not None:
            for ct in ["COFINSAliq","COFINSQtde","COFINSNT","COFINSOutr"]:
                cn = cof_node.find(f"nfe:{ct}", NS)
                if cn is not None:
                    v = safe_float(get_text(cn,"nfe:pCOFINS") or get_text(cn,"nfe:vAliqProd"))
                    if v > 0:
                        cof_vals.append(v)
                    break
    def majoritaria(vals):
        if not vals:
            return 0.0
        return max(set(vals), key=vals.count)
    return majoritaria(pis_vals), majoritaria(cof_vals)

# ─────────────────────────────────────────────
# ZIP
# ─────────────────────────────────────────────
def extrair_xmls_importacao_do_zip(zip_bytes: bytes) -> tuple:
    xmls_importacao, ignorados, erros_parse = [], [], []
    total_xml = 0
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            for entry in zf.infolist():
                nome = entry.filename
                if entry.is_dir() or not nome.lower().endswith(".xml"):
                    continue
                total_xml += 1
                xml_bytes_item = zf.read(nome)
                if not xml_bytes_item:
                    erros_parse.append(f"{nome.split('/')[-1]} (arquivo vazio no ZIP)")
                    continue
                root, erro = parse_xml_seguro(xml_bytes_item)
                if root is None:
                    erros_parse.append(f"{nome.split('/')[-1]} ({erro})")
                    continue
                nfe = root.find("nfe:NFe", NS) or root
                det_list = nfe.findall("nfe:infNFe/nfe:det", NS)
                cfop = ""
                if det_list:
                    cfop = get_text(det_list[0].find("nfe:prod", NS), "nfe:CFOP")
                if cfop.startswith("3"):
                    xmls_importacao.append({
                        "nome": nome.split("/")[-1],
                        "bytes": sanitizar_xml_bytes(xml_bytes_item)
                    })
                else:
                    ignorados.append(f"{nome.split('/')[-1]} (CFOP {cfop or 'N/A'})")
    except zipfile.BadZipFile:
        return [], 0, [], ["Arquivo ZIP invalido ou corrompido."]
    return xmls_importacao, total_xml, ignorados, erros_parse

# ─────────────────────────────────────────────
# DETECÇÃO
# ─────────────────────────────────────────────
def get_grupo_por_cfop(cfop: str) -> int:
    if not cfop:
        return 1
    mapa = {"11":2,"12":2,"13":3,"14":3,"15":9,"16":10,"17":11,"20":2,"21":2,"22":3,
            "25":9,"30":2,"31":3,"35":9,"40":2,"41":2,"55":12,"60":2}
    return mapa.get(cfop[:2], 1)

def get_grupo_por_ncm(ncm: str) -> int:
    if not ncm or len(ncm) < 2:
        return 1
    mapa = {"84":10,"85":10,"86":10,"87":10,"88":10,"89":10,"90":10,"91":10,"94":10,
            "28":3,"29":3,"30":3,"31":3,"32":3,"33":3,"34":3,"38":3,"39":3,"40":3,
            "44":3,"47":3,"48":3,"72":3,"73":3,"74":3,"75":3,"76":3,"82":3,"83":3,
            "01":2,"02":2,"03":2,"04":2,"07":2,"08":2,"09":2,"10":2,"16":2,"17":2,
            "18":2,"19":2,"20":2,"21":2,"22":2,"27":2}
    return mapa.get(ncm[:2], 1)

def detectar_grupo(cfop: str, ncm: str, grupo_padrao: int) -> int:
    if grupo_padrao > 0:
        return grupo_padrao
    g = get_grupo_por_cfop(cfop)
    if g == 1 and ncm:
        g2 = get_grupo_por_ncm(ncm)
        return g2 if g2 != 1 else g
    return g

def is_nota_importacao(nfe_root) -> bool:
    dest = nfe_root.find("nfe:infNFe/nfe:dest", NS)
    if dest is not None:
        ender = dest.find("nfe:enderDest", NS)
        if ender is not None and get_text(ender, "nfe:UF") == "EX":
            return True
    det_list = nfe_root.findall("nfe:infNFe/nfe:det", NS)
    if det_list:
        cfop = get_text(det_list[0].find("nfe:prod", NS), "nfe:CFOP")
        if cfop.startswith("3"):
            return True
    return False

def extrair_cnpj_empresa(nfe_root, cnpj_fallback: str) -> tuple:
    importacao = is_nota_importacao(nfe_root)
    if not importacao:
        dest = nfe_root.find("nfe:infNFe/nfe:dest", NS)
        if dest is not None:
            cnpj = get_text(dest, "nfe:CNPJ")
            if cnpj:
                return cnpj, "XML — <dest><CNPJ>"
            cpf = get_text(dest, "nfe:CPF")
            if cpf:
                return cpf, "XML — <dest><CPF>"
    if cnpj_fallback:
        return somente_numeros(cnpj_fallback), "Manual (fallback)"
    emit = nfe_root.find("nfe:infNFe/nfe:emit", NS)
    cnpj_emit = get_text(emit, "nfe:CNPJ") if emit is not None else ""
    if cnpj_emit:
        return cnpj_emit, "XML — <emit><CNPJ> (fallback)"
    return "", "Nao encontrado"

def detectar_ipi_zero_nao_isento(nfe_root) -> bool:
    CST_IPI_ISENTAS = {"02","03","04","05"}
    for det in nfe_root.findall("nfe:infNFe/nfe:det", NS):
        imp = det.find("nfe:imposto", NS)
        if imp is None:
            continue
        ipi_trib = imp.find("nfe:IPI/nfe:IPITrib", NS)
        if ipi_trib is not None:
            cst   = get_text(ipi_trib, "nfe:CST").strip().zfill(2)
            v_ipi = safe_float(get_text(ipi_trib, "nfe:vIPI"))
            if v_ipi == 0 and cst not in CST_IPI_ISENTAS:
                return True
    return False

# ─────────────────────────────────────────────
# EXCEL
# MUDANÇA 5 — normalizar CST em extrair_dados_impostos_itens
# ─────────────────────────────────────────────
def extrair_dados_impostos_itens(nfe_root, nome_arquivo: str,
                                  aliq_pis_pad: float = 0.0,
                                  aliq_cof_pad: float = 0.0,
                                  tem_direito_credito: bool = True) -> list:
    ide   = nfe_root.find("nfe:infNFe/nfe:ide", NS)
    emit  = nfe_root.find("nfe:infNFe/nfe:emit", NS)
    dest  = nfe_root.find("nfe:infNFe/nfe:dest", NS)
    importacao = is_nota_importacao(nfe_root)
    nNF   = get_text(ide, "nfe:nNF")
    dhEmi = fmt_date(get_text(ide, "nfe:dhEmi"))
    chave = extrair_chave_nfe(nfe_root)
    emit_nome = get_text(emit, "nfe:xNome") if emit is not None else ""
    forn_nome = get_text(dest, "nfe:xNome") if (importacao and dest is not None) else emit_nome
    linhas = []
    for det in nfe_root.findall("nfe:infNFe/nfe:det", NS):
        seq  = det.get("nItem", "")
        prod = det.find("nfe:prod", NS)
        imp  = det.find("nfe:imposto", NS)
        cod_prod  = get_text(prod, "nfe:cProd")
        desc_prod = get_text(prod, "nfe:xProd")
        ncm       = get_text(prod, "nfe:NCM")
        cfop      = get_text(prod, "nfe:CFOP")
        qtd       = get_text(prod, "nfe:qCom")
        v_unit    = get_text(prod, "nfe:vUnCom")
        v_prod    = get_text(prod, "nfe:vProd")
        v_outro   = get_text(prod, "nfe:vOutro")
        v_desc    = get_text(prod, "nfe:vDesc")
        cst_icms = aliq_icms = bc_icms = v_icms = v_icms_des = ""
        if imp is not None:
            for tp in ["ICMS00","ICMS10","ICMS20","ICMS30","ICMS40","ICMS51","ICMS60",
                       "ICMS70","ICMS90","ICMSSN101","ICMSSN102","ICMSSN201",
                       "ICMSSN202","ICMSSN500","ICMSSN900"]:
                node = imp.find(f"nfe:ICMS/nfe:{tp}", NS)
                if node is not None:
                    cst_icms  = get_text(node,"nfe:CST") or get_text(node,"nfe:CSOSN")
                    aliq_icms = get_text(node,"nfe:pICMS")
                    bc_icms   = get_text(node,"nfe:vBC")
                    v_icms    = get_text(node,"nfe:vICMS")
                    v_icms_des= get_text(node,"nfe:vICMSDeson")
                    break
        cst_ipi = aliq_ipi = bc_ipi = v_ipi = ""
        if imp is not None:
            ipi_trib = imp.find("nfe:IPI/nfe:IPITrib", NS)
            ipi_nt   = imp.find("nfe:IPI/nfe:IPINT", NS)
            if ipi_trib is not None:
                cst_ipi  = get_text(ipi_trib,"nfe:CST")
                aliq_ipi = get_text(ipi_trib,"nfe:pIPI")
                bc_ipi   = get_text(ipi_trib,"nfe:vBC")
                v_ipi    = get_text(ipi_trib,"nfe:vIPI")
            elif ipi_nt is not None:
                cst_ipi = get_text(ipi_nt,"nfe:CST")
                v_ipi   = "0.00"

        # ── MUDANÇA 5: normalizar CST PIS/COFINS com zfill(2) ──
        cst_pis_xml = aliq_pis = bc_pis = v_pis = ""
        if imp is not None:
            pis_node = imp.find("nfe:PIS", NS)
            if pis_node is not None:
                for pt in ["PISAliq","PISQtde","PISNT","PISOutr"]:
                    pn = pis_node.find(f"nfe:{pt}", NS)
                    if pn is not None:
                        # NORMALIZAÇÃO: zfill(2) garante "01" em vez de "1"
                        cst_pis_xml = str(get_text(pn,"nfe:CST")).strip().zfill(2)
                        aliq_pis    = get_text(pn,"nfe:pPIS") or get_text(pn,"nfe:vAliqProd")
                        bc_pis      = get_text(pn,"nfe:vBC")
                        v_pis       = get_text(pn,"nfe:vPIS")
                        break
        cst_cof_xml = aliq_cof = bc_cof = v_cof = ""
        if imp is not None:
            cof_node = imp.find("nfe:COFINS", NS)
            if cof_node is not None:
                for ct in ["COFINSAliq","COFINSQtde","COFINSNT","COFINSOutr"]:
                    cn = cof_node.find(f"nfe:{ct}", NS)
                    if cn is not None:
                        # NORMALIZAÇÃO: zfill(2)
                        cst_cof_xml = str(get_text(cn,"nfe:CST")).strip().zfill(2)
                        aliq_cof    = get_text(cn,"nfe:pCOFINS") or get_text(cn,"nfe:vAliqProd")
                        bc_cof      = get_text(cn,"nfe:vBC")
                        v_cof       = get_text(cn,"nfe:vCOFINS")
                        break

        bc_ii = v_ii = ""
        if imp is not None:
            ii_node = imp.find("nfe:II", NS)
            if ii_node is not None:
                bc_ii = get_text(ii_node,"nfe:vBC")
                v_ii  = get_text(ii_node,"nfe:vII")

        aliq_pis_f = safe_float(aliq_pis)
        aliq_cof_f = safe_float(aliq_cof)

        # MUDANÇA 5: usa obter_cst_entrada em vez das funções legado
        cst_pis_ef = obter_cst_entrada(cst_pis_xml, tem_direito_credito)
        cst_cof_ef = obter_cst_entrada(cst_cof_xml, tem_direito_credito)

        def _f(v): return safe_float(v)
        linhas.append({
            "Arquivo":nome_arquivo,"NF":nNF,"Emissao":dhEmi,"Fornecedor":forn_nome,
            "Chave NF-e":chave,"Item":seq,"Cod. Produto":cod_prod,"Descricao":desc_prod,
            "NCM":ncm,"CFOP":cfop,"Qtd":_f(qtd),"V. Unit.":_f(v_unit),"V. Prod.":_f(v_prod),
            "V. Outro":_f(v_outro),"V. Desc.":_f(v_desc),"CST ICMS":cst_icms,
            "BC ICMS":_f(bc_icms),"Aliq. ICMS %":_f(aliq_icms),"V. ICMS":_f(v_icms),
            "V. ICMS Deson":_f(v_icms_des),"CST IPI":cst_ipi,"BC IPI":_f(bc_ipi),
            "Aliq. IPI %":_f(aliq_ipi),"V. IPI":_f(v_ipi),"BC II":_f(bc_ii),"V. II":_f(v_ii),
            "CST PIS (XML)":cst_pis_xml,"CST PIS (Efet)":cst_pis_ef,"BC PIS":_f(bc_pis),
            "Aliq. PIS %":aliq_pis_f,"V. PIS":_f(v_pis),"CST COF (XML)":cst_cof_xml,
            "CST COF (Efet)":cst_cof_ef,"BC COFINS":_f(bc_cof),"Aliq. COFINS %":aliq_cof_f,
            "V. COFINS":_f(v_cof),"Aliq. PIS Padrao":aliq_pis_pad,"Aliq. COF Padrao":aliq_cof_pad,
            "PIS c/ Credito":"SIM" if cst_pis_ef != cst_pis_xml else "NAO",
            "COF c/ Credito":"SIM" if cst_cof_ef != cst_cof_xml else "NAO",
        })
    return linhas

def gerar_excel_relatorio(dados_itens: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Impostos por Item"
    cinza_esc="444444"; laranja="FF8000"
    subhdr_icms=PatternFill("solid",fgColor="1565C0"); subhdr_ipi=PatternFill("solid",fgColor="6A1B9A")
    subhdr_ii=PatternFill("solid",fgColor="BF360C");   subhdr_pis=PatternFill("solid",fgColor="1B5E20")
    subhdr_cof=PatternFill("solid",fgColor="E65100");  subhdr_prod=PatternFill("solid",fgColor=laranja)
    subhdr_ident=PatternFill("solid",fgColor=cinza_esc); subhdr_red=PatternFill("solid",fgColor="880000")
    white_font=Font(bold=True,color="FFFFFF",size=9)
    alt_fill=PatternFill("solid",fgColor="E9E9E9"); alt_fill2=PatternFill("solid",fgColor="FFFFFF")
    center=Alignment(horizontal="center",vertical="center"); left=Alignment(horizontal="left",vertical="center")
    thin=Side(style="thin",color="CCCCCC"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    green_fill=PatternFill("solid",fgColor="E8F5E9")
    if not dados_itens:
        ws["A1"]="Nenhum dado disponivel."
        buf=io.BytesIO(); wb.save(buf); return buf.getvalue()
    colunas=list(dados_itens[0].keys())
    grupos={
        "Identificacao":["Arquivo","NF","Emissao","Fornecedor","Chave NF-e","Item"],
        "Produto":["Cod. Produto","Descricao","NCM","CFOP","Qtd","V. Unit.","V. Prod.","V. Outro","V. Desc."],
        "ICMS":["CST ICMS","BC ICMS","Aliq. ICMS %","V. ICMS","V. ICMS Deson"],
        "IPI":["CST IPI","BC IPI","Aliq. IPI %","V. IPI"],
        "II":["BC II","V. II"],
        "PIS":["CST PIS (XML)","CST PIS (Efet)","BC PIS","Aliq. PIS %","V. PIS"],
        "COFINS":["CST COF (XML)","CST COF (Efet)","BC COFINS","Aliq. COFINS %","V. COFINS"],
        "Credito":["Aliq. PIS Padrao","Aliq. COF Padrao","PIS c/ Credito","COF c/ Credito"],
    }
    cor_grupo={
        "Identificacao":subhdr_ident,"Produto":subhdr_prod,"ICMS":subhdr_icms,
        "IPI":subhdr_ipi,"II":subhdr_ii,"PIS":subhdr_pis,"COFINS":subhdr_cof,
        "Credito":subhdr_red
    }
    col_grupo={}
    for grp,cols in grupos.items():
        for c in cols:
            col_grupo[c]=grp
    col_idx=1
    for grp,cols in grupos.items():
        start=col_idx; end=col_idx+len(cols)-1
        ws.merge_cells(start_row=1,start_column=start,end_row=1,end_column=end)
        cell=ws.cell(row=1,column=start,value=grp)
        cell.fill=cor_grupo[grp]; cell.font=white_font; cell.alignment=center; cell.border=border
        col_idx=end+1
    for ci,col in enumerate(colunas,start=1):
        cell=ws.cell(row=2,column=ci,value=col)
        grp=col_grupo.get(col,"Identificacao")
        cell.fill=cor_grupo[grp]; cell.font=white_font; cell.alignment=center; cell.border=border
    cols_num={"Qtd","V. Unit.","V. Prod.","V. Outro","V. Desc.","BC ICMS","Aliq. ICMS %","V. ICMS",
              "V. ICMS Deson","BC IPI","Aliq. IPI %","V. IPI","BC II","V. II","BC PIS","Aliq. PIS %",
              "V. PIS","BC COFINS","Aliq. COFINS %","V. COFINS","Aliq. PIS Padrao","Aliq. COF Padrao"}
    for ri,row in enumerate(dados_itens,start=3):
        fill_base=alt_fill if ri%2==1 else alt_fill2
        pis_cred=row.get("PIS c/ Credito")=="SIM"
        cof_cred=row.get("COF c/ Credito")=="SIM"
        for ci,col in enumerate(colunas,start=1):
            val=row[col]; cell=ws.cell(row=ri,column=ci,value=val)
            cell.border=border; cell.alignment=center if col in cols_num else left
            if col in ("CST PIS (Efet)","BC PIS","Aliq. PIS %","V. PIS","CST PIS (XML)") and pis_cred:
                cell.fill=green_fill
            elif col in ("CST COF (Efet)","BC COFINS","Aliq. COFINS %","V. COFINS","CST COF (XML)") and cof_cred:
                cell.fill=green_fill
            elif col in ("PIS c/ Credito","COF c/ Credito") and val=="SIM":
                cell.fill=green_fill
            else:
                cell.fill=fill_base
            if col in cols_num and isinstance(val,float):
                cell.number_format='#,##0.0000' if ("Aliq" in col or "Padrao" in col) else '#,##0.00'
    tot_row=len(dados_itens)+3
    ws.cell(row=tot_row,column=1,value="TOTAL").font=Font(bold=True)
    cols_soma=["V. Prod.","V. Outro","V. Desc.","BC ICMS","V. ICMS","V. ICMS Deson",
               "BC IPI","V. IPI","BC II","V. II","BC PIS","V. PIS","BC COFINS","V. COFINS"]
    tot_fill=PatternFill("solid",fgColor="FFF3E0")
    for ci,col in enumerate(colunas,start=1):
        cell=ws.cell(row=tot_row,column=ci); cell.fill=tot_fill; cell.border=border
        cell.font=Font(bold=True,size=9)
        if col in cols_soma:
            cell.value=sum(r[col] for r in dados_itens if isinstance(r[col],float))
            cell.number_format='#,##0.00'; cell.alignment=center
    larguras={
        "Arquivo":22,"NF":10,"Emissao":12,"Fornecedor":30,"Chave NF-e":46,"Item":6,
        "Cod. Produto":18,"Descricao":50,"NCM":12,"CFOP":8,"Qtd":10,"V. Unit.":14,
        "V. Prod.":14,"V. Outro":12,"V. Desc.":12,"CST ICMS":10,"BC ICMS":14,
        "Aliq. ICMS %":12,"V. ICMS":14,"V. ICMS Deson":14,"CST IPI":10,"BC IPI":14,
        "Aliq. IPI %":12,"V. IPI":12,"BC II":14,"V. II":12,"CST PIS (XML)":13,
        "CST PIS (Efet)":13,"BC PIS":14,"Aliq. PIS %":12,"V. PIS":12,"CST COF (XML)":13,
        "CST COF (Efet)":13,"BC COFINS":14,"Aliq. COFINS %":14,"V. COFINS":12,
        "Aliq. PIS Padrao":16,"Aliq. COF Padrao":16,"PIS c/ Credito":14,"COF c/ Credito":14
    }
    for ci,col in enumerate(colunas,start=1):
        ws.column_dimensions[get_column_letter(ci)].width=larguras.get(col,14)
    ws.row_dimensions[1].height=20; ws.row_dimensions[2].height=28; ws.freeze_panes="A3"
    ws2=wb.create_sheet("Resumo 1020")
    ws2["A1"]="Resumo das linhas 1020 (agrupado por aliquota)"
    ws2["A1"].font=Font(bold=True,color=laranja,size=11)
    hdr2=["Sigla","Base de Calculo","Aliquota %","Valor Imposto","CST Efetivo","Obs"]
    for ci,h in enumerate(hdr2,start=1):
        c=ws2.cell(row=2,column=ci,value=h)
        c.fill=PatternFill("solid",fgColor=cinza_esc); c.font=white_font
        c.alignment=center; c.border=border
    pis_ag={}; cof_ag={}
    for row in dados_itens:
        k=row["Aliq. PIS %"]
        if k not in pis_ag:
            pis_ag[k]={"bc":0.0,"val":0.0,"cst":row["CST PIS (Efet)"]}
        pis_ag[k]["bc"]+=row["BC PIS"]; pis_ag[k]["val"]+=row["V. PIS"]
        k2=row["Aliq. COFINS %"]
        if k2 not in cof_ag:
            cof_ag[k2]={"bc":0.0,"val":0.0,"cst":row["CST COF (Efet)"]}
        cof_ag[k2]["bc"]+=row["BC COFINS"]; cof_ag[k2]["val"]+=row["V. COFINS"]
    ri2=3
    for aliq,d in sorted(pis_ag.items()):
        if d["val"]>0 or d["bc"]>0:
            cst_desc = TABELA_CST_PIS_COFINS_ENTRADA.get(d["cst"], d["cst"])
            row_fill=green_fill if d["cst"] in ("50","51","52") else alt_fill2
            for ci,v in enumerate(["PIS",d["bc"],aliq,d["val"],d["cst"],cst_desc],start=1):
                c=ws2.cell(row=ri2,column=ci,value=v); c.border=border
                c.alignment=center; c.fill=row_fill
                if ci in (2,3,4):
                    c.number_format='#,##0.0000' if ci==3 else '#,##0.00'
            ri2+=1
    for aliq,d in sorted(cof_ag.items()):
        if d["val"]>0 or d["bc"]>0:
            cst_desc = TABELA_CST_PIS_COFINS_ENTRADA.get(d["cst"], d["cst"])
            row_fill=green_fill if d["cst"] in ("50","51","52") else alt_fill2
            for ci,v in enumerate(["COFINS",d["bc"],aliq,d["val"],d["cst"],cst_desc],start=1):
                c=ws2.cell(row=ri2,column=ci,value=v); c.border=border
                c.alignment=center; c.fill=row_fill
                if ci in (2,3,4):
                    c.number_format='#,##0.0000' if ci==3 else '#,##0.00'
            ri2+=1
    for ci in range(1,7):
        ws2.column_dimensions[get_column_letter(ci)].width=22
    ws2.column_dimensions[get_column_letter(6)].width=60
    ws2.freeze_panes="A3"
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

# ─────────────────────────────────────────────
# REGISTROS 0000 / 0020 / 0100 / 0110
# ─────────────────────────────────────────────
def gerar_registro_0000(cnpj_empresa: str) -> str:
    return pipe_join(["0000", cnpj_empresa])

def gerar_registro_0020(emit, dest=None, is_importacao: bool = False) -> str:
    if is_importacao and dest is not None:
        razao       = get_text(dest,"nfe:xNome")[:150].upper()
        fantasia    = razao[:40]
        ender       = dest.find("nfe:enderDest", NS)
        logradouro  = get_text(ender,"nfe:xLgr")                  if ender is not None else ""
        numero      = somente_numeros(get_text(ender,"nfe:nro"))   if ender is not None else ""
        complemento = ""
        bairro      = get_text(ender,"nfe:xBairro")               if ender is not None else ""
        cod_mun     = somente_numeros(get_text(ender,"nfe:cMun"))  if ender is not None else ""
        cep         = get_text(ender,"nfe:CEP")                   if ender is not None else ""
        c_pais_xml  = get_text(ender,"nfe:cPais")                 if ender is not None else ""
        x_pais_xml  = get_text(ender,"nfe:xPais")                 if ender is not None else ""
        cod_pais    = resolver_codigo_pais_dominio(c_pais_xml, x_pais_xml)
        inscricao   = ""; uf_campo = "EX"; ie = ""; regime = "N"; contrib = "N"
    else:
        inscricao    = get_text(emit,"nfe:CNPJ")
        razao        = get_text(emit,"nfe:xNome")[:150].upper()
        fantasia_raw = get_text(emit,"nfe:xFant")
        fantasia     = fantasia_raw[:40].upper() if fantasia_raw else razao[:40]
        ender        = emit.find("nfe:enderEmit", NS)
        logradouro   = get_text(ender,"nfe:xLgr")                 if ender is not None else ""
        numero       = somente_numeros(get_text(ender,"nfe:nro"))  if ender is not None else ""
        complemento  = get_text(ender,"nfe:xCpl")                 if ender is not None else ""
        bairro       = get_text(ender,"nfe:xBairro")              if ender is not None else ""
        cod_mun      = somente_numeros(get_text(ender,"nfe:cMun")) if ender is not None else ""
        cep          = get_text(ender,"nfe:CEP")                  if ender is not None else ""
        uf_campo     = get_text(ender,"nfe:UF")                   if ender is not None else ""
        cod_pais     = ""
        ie           = get_text(emit,"nfe:IE")
        crt          = get_text(emit,"nfe:CRT")
        regime       = {"1":"M","2":"E","3":"N"}.get(crt,"N")
        contrib      = "S" if ie and ie.upper() not in ("ISENTO","NAO CONTRIBUINTE","") else "N"
    c = [""] * 33
    c[0]="0020"; c[1]=inscricao; c[2]=razao; c[3]=fantasia
    c[4]=logradouro; c[5]=numero; c[6]=complemento; c[7]=bairro
    c[8]=cod_mun; c[9]=uf_campo; c[10]=cod_pais; c[11]=cep
    c[12]=ie; c[13]=""; c[14]=""; c[15]=""
    c[16]=""; c[17]=""; c[18]=""; c[19]=""
    c[20]=""; c[21]="N"; c[22]="7"; c[23]=regime
    c[24]=contrib; c[25]=""; c[26]=""; c[27]=""
    c[28]=""; c[29]="N"; c[30]="N"; c[31]=""
    c[32]=""
    return pipe_join(c)

def gerar_registro_0100(det, grupo_padrao: int = 0) -> str:
    prod      = det.find("nfe:prod", NS)
    cod_prod  = get_text(prod,"nfe:cProd")[:14]
    descricao = get_text(prod,"nfe:xProd").upper()
    ncm       = get_text(prod,"nfe:NCM")
    unidade   = get_text(prod,"nfe:uCom")
    val_unit  = get_text(prod,"nfe:vUnCom")
    cest      = get_text(prod,"nfe:CEST")
    cfop      = get_text(prod,"nfe:CFOP")
    cod_grupo = detectar_grupo(cfop, ncm, grupo_padrao)
    imposto   = det.find("nfe:imposto", NS)
    cst_icms = aliq_icms = aliq_ipi = ""
    if imposto is not None:
        for tp in ["ICMS00","ICMS10","ICMS20","ICMS30","ICMS40","ICMS51","ICMS60",
                   "ICMS70","ICMS90","ICMSSN101","ICMSSN102","ICMSSN201",
                   "ICMSSN202","ICMSSN500","ICMSSN900"]:
            node = imposto.find(f"nfe:ICMS/nfe:{tp}", NS)
            if node is not None:
                cst_icms  = get_text(node,"nfe:CST") or get_text(node,"nfe:CSOSN")
                aliq_icms = fmt_decimal(get_text(node,"nfe:pICMS"))
                break
        ipi_trib = imposto.find("nfe:IPI/nfe:IPITrib", NS)
        if ipi_trib is not None:
            aliq_ipi = fmt_decimal(get_text(ipi_trib,"nfe:pIPI"))
    c = [""] * 91
    c[0]="0100"; c[1]=cod_prod; c[2]=descricao; c[3]=""; c[4]=ncm; c[5]=""; c[6]=""; c[7]=""
    c[8]=str(cod_grupo); c[9]=unidade; c[10]="N"; c[11]="O"; c[12]=""; c[13]=""; c[14]=""; c[15]="N"
    c[16]=""; c[17]=fmt_decimal(val_unit,3); c[18]=""; c[19]=""; c[20]=cst_icms; c[21]=aliq_icms
    c[22]=aliq_ipi; c[23]="M"; c[24]=""; c[25]="N"
    for i in range(26,74):
        c[i]=""
    c[74]=DATA_CADASTRO_FIXO
    for i in range(75,88):
        c[i]=""
    c[88]=cest; c[89]=""; c[90]=""
    return pipe_join(c)

# ─────────────────────────────────────────────
# MUDANÇA 3 — extrair_pis_cofins: normalizar CST + propagar tem_direito_credito
# ─────────────────────────────────────────────
def extrair_pis_cofins(det,
                        importacao: bool = False,
                        aliq_pis_pad: float = 0.0,
                        aliq_cof_pad: float = 0.0,
                        tem_direito_credito: bool = True) -> dict:
    """
    Extrai CST, alíquotas e BC de PIS/COFINS de um item.
    - Normaliza o CST com zfill(2) antes de qualquer lógica.
    - Usa obter_cst_entrada para determinar o CST de entrada efetivo.
    - cst_s é o CST de saída equivalente ao CST de entrada efetivo.
    """
    imposto = det.find("nfe:imposto", NS)
    res = {
        "cst_e":       "",
        "aliq_pis_e":  "",
        "aliq_cof_e":  "",
        "cst_s":       "",
        "aliq_pis_s":  "",
        "aliq_cof_s":  "",
        "class_trib":  "",
        "cst_cof_e":   "",
    }
    if imposto is None:
        return res

    pis_node = imposto.find("nfe:PIS", NS)
    if pis_node is not None:
        for pt in ["PISAliq","PISQtde","PISNT","PISOutr"]:
            pn = pis_node.find(f"nfe:{pt}", NS)
            if pn is not None:
                # MUDANÇA 3: normalizar CST com zfill(2)
                cst_xml        = str(get_text(pn,"nfe:CST")).strip().zfill(2)
                aliq           = get_text(pn,"nfe:pPIS") or get_text(pn,"nfe:vAliqProd")
                res["cst_e"]      = obter_cst_entrada(cst_xml, tem_direito_credito)
                res["aliq_pis_e"] = fmt_decimal(aliq, 4)
                break

    cof_node = imposto.find("nfe:COFINS", NS)
    if cof_node is not None:
        for ct in ["COFINSAliq","COFINSQtde","COFINSNT","COFINSOutr"]:
            cn = cof_node.find(f"nfe:{ct}", NS)
            if cn is not None:
                # MUDANÇA 3: normalizar CST com zfill(2)
                cst_xml        = str(get_text(cn,"nfe:CST")).strip().zfill(2)
                aliq           = get_text(cn,"nfe:pCOFINS") or get_text(cn,"nfe:vAliqProd")
                res["aliq_cof_e"] = fmt_decimal(aliq, 4)
                res["cst_cof_e"]  = obter_cst_entrada(cst_xml, tem_direito_credito)
                break

    # cst_s: CST de saída correspondente ao CST de entrada efetivo
    res["cst_s"]      = CST_ENTRADA_SAIDA.get(res["cst_e"], "")
    res["aliq_pis_s"] = res["aliq_pis_e"]
    res["aliq_cof_s"] = res["aliq_cof_e"]

    ibs_node = imposto.find("nfe:IBSCBS", NS)
    if ibs_node is not None:
        res["class_trib"] = get_text(ibs_node,"nfe:cClassTrib")

    return res

# ─────────────────────────────────────────────
# MUDANÇA 3 (continuação) — gerar_registro_0110: propagar tem_direito_credito
# ─────────────────────────────────────────────
def gerar_registro_0110(det,
                         importacao: bool = False,
                         aliq_pis_pad: float = 0.0,
                         aliq_cof_pad: float = 0.0,
                         tem_direito_credito: bool = True) -> str:
    pc = extrair_pis_cofins(
        det,
        importacao=importacao,
        aliq_pis_pad=aliq_pis_pad,
        aliq_cof_pad=aliq_cof_pad,
        tem_direito_credito=tem_direito_credito,
    )
    ct = pc["class_trib"]
    vinculo_credito = "08" if importacao else ""
    c = [""] * 70
    c[0]="0110"; c[1]=DATA_CADASTRO_FIXO; c[2]=pc["cst_e"]; c[3]=vinculo_credito
    c[4]="01"; c[5]="N"; c[6]="N"; c[7]=pc["aliq_pis_e"]; c[8]=pc["aliq_cof_e"]
    c[9]="N"; c[10]="N"; c[11]=""; c[12]=""; c[13]=""; c[14]=""
    c[15]=pc["cst_s"]; c[16]="N"; c[17]=""; c[18]=""; c[19]=""
    c[20]="N"; c[21]=pc["aliq_pis_s"]; c[22]=pc["aliq_cof_s"]
    c[23]="N"; c[24]="N"
    for i in range(25, 40):
        c[i]=""
    c[40]="N"; c[41]="N"; c[42]="N"; c[43]=""; c[44]="N"; c[45]=""; c[46]="N"
    for i in range(47, 66):
        c[i]=""
    c[66]=ct; c[67]=ct; c[68]="N"; c[69]="N"
    return pipe_join(c)

def gerar_registro_1000(nfe_root, cnpj_empresa: str,
                        acumulador: str = "1157",
                        especie: str = "36",
                        importacao: bool = False) -> str:
    ide   = nfe_root.find("nfe:infNFe/nfe:ide", NS)
    emit  = nfe_root.find("nfe:infNFe/nfe:emit", NS)
    dest  = nfe_root.find("nfe:infNFe/nfe:dest", NS)
    total = nfe_root.find("nfe:infNFe/nfe:total/nfe:ICMSTot", NS)
    if importacao and dest is not None:
        id_ext    = get_text(dest,"nfe:idEstrangeiro").strip()
        cnpj_forn = id_ext if id_ext else ""
    else:
        cnpj_forn = get_text(emit,"nfe:CNPJ")
    emitente_nf = "P" if importacao else "T"
    ie_forn     = "" if importacao else get_text(emit,"nfe:IE")
    nNF      = get_text(ide,"nfe:nNF")
    serie    = get_text(ide,"nfe:serie")
    dhEmi    = fmt_date(get_text(ide,"nfe:dhEmi"))
    c_mun_fg = get_text(ide,"nfe:cMunFG")
    det_list   = nfe_root.findall("nfe:infNFe/nfe:det", NS)
    cfop_first = ""
    if det_list:
        cfop_first = get_text(det_list[0].find("nfe:prod", NS),"nfe:CFOP")
    v_nf     = fmt_decimal(get_text(total,"nfe:vNF"))
    v_pis    = fmt_decimal(get_text(total,"nfe:vPIS"))
    v_cofins = fmt_decimal(get_text(total,"nfe:vCOFINS"))
    v_ipi    = fmt_decimal(get_text(total,"nfe:vIPI"))
    v_st     = fmt_decimal(get_text(total,"nfe:vST"))
    v_prod   = fmt_decimal(get_text(total,"nfe:vProd"))
    v_frete  = fmt_decimal(get_text(total,"nfe:vFrete"))
    v_seg    = fmt_decimal(get_text(total,"nfe:vSeg"))
    v_outro  = fmt_decimal(get_text(total,"nfe:vOutro"))
    v_icms_d = fmt_decimal(get_text(total,"nfe:vICMSDeson"))
    chave    = extrair_chave_nfe(nfe_root)
    transp        = nfe_root.find("nfe:infNFe/nfe:transp", NS)
    mod_frete_cod = get_text(transp,"nfe:modFrete")
    mod_frete     = {"0":"C","1":"F","2":"S","3":"T","4":"R","5":"D","9":"S"}.get(mod_frete_cod,"C")
    inf_adic  = nfe_root.find("nfe:infNFe/nfe:infAdic", NS)
    obs_fisco = ""
    if inf_adic is not None:
        obs_fisco = get_text(inf_adic,"nfe:infAdFisco")[:300]
    n_di = ""
    if det_list:
        di_node = det_list[0].find("nfe:prod/nfe:DI", NS)
        if di_node is not None:
            n_di = get_text(di_node,"nfe:nDI")
    c = [""] * 98
    c[0]="1000"; c[1]=especie; c[2]=cnpj_forn; c[3]=""; c[4]=acumulador
    c[5]=cfop_first; c[6]=""; c[7]=nNF; c[8]=serie; c[9]=""
    c[10]=dhEmi; c[11]=dhEmi; c[12]=v_nf; c[13]=""; c[14]=obs_fisco
    c[15]=mod_frete; c[16]=emitente_nf; c[17]=""; c[18]=""; c[19]=""
    c[20]=""; c[21]=""; c[22]=""; c[23]=""; c[24]=""
    c[25]=v_frete; c[26]=v_seg; c[27]=v_outro; c[28]=v_pis; c[29]=""
    c[30]=v_cofins; c[31]=""; c[32]=""; c[33]=""; c[34]=""
    c[35]=""; c[36]=""; c[37]=""; c[38]=v_prod; c[39]=c_mun_fg
    c[40]="0"; c[41]=""; c[42]=""; c[43]=ie_forn; c[44]=""
    c[45]=""; c[46]=""; c[47]=""; c[48]=""; c[49]=""
    c[50]=""; c[51]=n_di; c[52]="N"; c[53]=chave; c[54]=""
    c[55]=""; c[56]=""; c[57]=""; c[58]=""
    c[59]="1" if importacao else ""
    c[60]=""; c[61]=""; c[62]=""; c[63]=""; c[64]=""
    c[65]=""; c[66]=""; c[67]=""; c[68]=""
    c[69]="1" if importacao else ""
    c[70]=""; c[71]=""; c[72]=""; c[73]=""; c[74]=""
    c[75]=""; c[76]=""; c[77]=""; c[78]=""; c[79]=""
    c[80]=""; c[81]=""; c[82]=""; c[83]=""; c[84]=""
    c[85]=""; c[86]=""; c[87]=""; c[88]=""
    c[89]=v_ipi; c[90]=v_st; c[91]=""; c[92]=""; c[93]=""
    c[94]=""; c[95]=""; c[96]=v_icms_d; c[97]=""
    return pipe_join(c)

def gerar_registros_1010(nfe_root) -> list:
    linhas = []
    inf_adic = nfe_root.find("nfe:infNFe/nfe:infAdic", NS)
    if inf_adic is None:
        return linhas
    for txt, cod in [(get_text(inf_adic,"nfe:infAdFisco"),"1"),
                     (get_text(inf_adic,"nfe:infCpl"),"2")]:
        if txt:
            for bloco in [txt[i:i+300] for i in range(0,len(txt),300)]:
                linhas.append(pipe_join(["1010", cod, bloco]))
    return linhas

def gerar_registros_1015(nfe_root) -> list:
    linhas = []
    inf_adic = nfe_root.find("nfe:infNFe/nfe:infAdic", NS)
    if inf_adic is None:
        return linhas
    for txt, cod in [(get_text(inf_adic,"nfe:infAdFisco"),"1"),
                     (get_text(inf_adic,"nfe:infCpl"),"2")]:
        if txt:
            for bloco in [txt[i:i+300] for i in range(0,len(txt),300)]:
                linhas.append(pipe_join(["1015", cod, bloco]))
    return linhas

def gerar_registros_1020(nfe_root, importacao: bool = False) -> list:
    total    = nfe_root.find("nfe:infNFe/nfe:total/nfe:ICMSTot", NS)
    det_list = nfe_root.findall("nfe:infNFe/nfe:det", NS)
    v_nf     = fmt_decimal(get_text(total,"nfe:vNF"))
    linhas   = []

    def r1020(cod, perc_red="", base="", aliq="", valor="",
              isentas="", outras="", v_ipi_r="", v_st_r="",
              v_cont="", cod_rec="", nao_trib="", parc_red="",
              aliq_int="", nat_rend="", tipo_ded="", tipo_isen="", descricao=""):
        return pipe_join(["1020", cod, perc_red, base, aliq, valor,
                          isentas, outras, v_ipi_r, v_st_r, v_cont, cod_rec,
                          nao_trib, parc_red, aliq_int, nat_rend, tipo_ded, tipo_isen, descricao])

    icms_por_aliq = {}
    for det in det_list:
        imp = det.find("nfe:imposto", NS)
        if imp is None:
            continue
        for tp in ["ICMS00","ICMS10","ICMS20","ICMS30","ICMS40","ICMS51","ICMS60",
                   "ICMS70","ICMS90","ICMSSN101","ICMSSN102","ICMSSN201",
                   "ICMSSN202","ICMSSN500","ICMSSN900"]:
            node = imp.find(f"nfe:ICMS/nfe:{tp}", NS)
            if node is not None:
                aliq_str = get_text(node,"nfe:pICMS") or "0"
                bc    = safe_float(get_text(node,"nfe:vBC"))
                valor = safe_float(get_text(node,"nfe:vICMS"))
                deson = safe_float(get_text(node,"nfe:vICMSDeson"))
                if aliq_str not in icms_por_aliq:
                    icms_por_aliq[aliq_str] = {"bc":0.0,"valor":0.0,"deson":0.0}
                icms_por_aliq[aliq_str]["bc"]    += bc
                icms_por_aliq[aliq_str]["valor"] += valor
                icms_por_aliq[aliq_str]["deson"] += deson
                break

    v_ipi_tot  = fmt_decimal(get_text(total,"nfe:vIPI"))
    v_st_tot   = fmt_decimal(get_text(total,"nfe:vST"))
    v_prod_tot = fmt_decimal(get_text(total,"nfe:vProd"))
    if importacao:
        if safe_float(v_ipi_tot) > 0:
            outras_icms = v_ipi_tot
        elif detectar_ipi_zero_nao_isento(nfe_root):
            outras_icms = v_prod_tot
        else:
            outras_icms = ""
    else:
        outras_icms = ""

    for aliq_str, dados in sorted(icms_por_aliq.items(), key=lambda x: safe_float(x[0])):
        if dados["valor"] > 0 or dados["bc"] > 0:
            linhas.append(r1020(1,
                base  = fmt_decimal(str(dados["bc"])),
                aliq  = fmt_decimal(aliq_str),
                valor = fmt_decimal(str(dados["valor"])),
                outras= outras_icms,
                v_ipi_r="" if outras_icms else v_ipi_tot,
                v_st_r=v_st_tot, v_cont=v_nf))

    ipi_por_aliq = {}
    for det in det_list:
        imp = det.find("nfe:imposto", NS)
        if imp is None:
            continue
        ipi_trib = imp.find("nfe:IPI/nfe:IPITrib", NS)
        ipi_nt   = imp.find("nfe:IPI/nfe:IPINT", NS)
        if ipi_trib is not None:
            aliq_str = get_text(ipi_trib,"nfe:pIPI") or "0"
            bc    = safe_float(get_text(ipi_trib,"nfe:vBC"))
            valor = safe_float(get_text(ipi_trib,"nfe:vIPI"))
            if aliq_str not in ipi_por_aliq:
                ipi_por_aliq[aliq_str] = {"bc":0.0,"valor":0.0,"isentas":0.0}
            ipi_por_aliq[aliq_str]["bc"]    += bc
            ipi_por_aliq[aliq_str]["valor"] += valor
        elif ipi_nt is not None:
            v_prod_item = safe_float(get_text(det.find("nfe:prod",NS),"nfe:vProd"))
            if "0" not in ipi_por_aliq:
                ipi_por_aliq["0"] = {"bc":0.0,"valor":0.0,"isentas":0.0}
            ipi_por_aliq["0"]["bc"]      += v_prod_item
            ipi_por_aliq["0"]["isentas"] += v_prod_item

    for aliq_str, dados in sorted(ipi_por_aliq.items(), key=lambda x: safe_float(x[0])):
        aliq_f = safe_float(aliq_str)
        if aliq_f == 0 and dados["isentas"] > 0:
            linhas.append(r1020(2, base=fmt_decimal(str(dados["bc"])),
                                isentas=fmt_decimal(str(dados["isentas"])), v_cont=v_nf))
        elif dados["valor"] > 0 or (dados["bc"] > 0 and aliq_f > 0):
            linhas.append(r1020(2, base=fmt_decimal(str(dados["bc"])),
                                aliq=fmt_decimal(aliq_str),
                                valor=fmt_decimal(str(dados["valor"])), v_cont=v_nf))

    pis_por_aliq = {}
    for det in det_list:
        imp = det.find("nfe:imposto", NS)
        if imp is None:
            continue
        pis_node = imp.find("nfe:PIS", NS)
        if pis_node is not None:
            for pt in ["PISAliq","PISQtde","PISNT","PISOutr"]:
                pn = pis_node.find(f"nfe:{pt}", NS)
                if pn is not None:
                    aliq_str = get_text(pn,"nfe:pPIS") or get_text(pn,"nfe:vAliqProd") or "0"
                    bc    = safe_float(get_text(pn,"nfe:vBC"))
                    valor = safe_float(get_text(pn,"nfe:vPIS"))
                    if aliq_str not in pis_por_aliq:
                        pis_por_aliq[aliq_str] = {"bc":0.0,"valor":0.0}
                    pis_por_aliq[aliq_str]["bc"]    += bc
                    pis_por_aliq[aliq_str]["valor"] += valor
                    break

    for aliq_str, dados in sorted(pis_por_aliq.items(), key=lambda x: safe_float(x[0])):
        if dados["valor"] > 0 or dados["bc"] > 0:
            linhas.append(r1020(4, base=fmt_decimal(str(dados["bc"])),
                                aliq=fmt_decimal(aliq_str,4),
                                valor=fmt_decimal(str(dados["valor"])), v_cont=v_nf))

    cof_por_aliq = {}
    for det in det_list:
        imp = det.find("nfe:imposto", NS)
        if imp is None:
            continue
        cof_node = imp.find("nfe:COFINS", NS)
        if cof_node is not None:
            for ct in ["COFINSAliq","COFINSQtde","COFINSNT","COFINSOutr"]:
                cn = cof_node.find(f"nfe:{ct}", NS)
                if cn is not None:
                    aliq_str = get_text(cn,"nfe:pCOFINS") or get_text(cn,"nfe:vAliqProd") or "0"
                    bc    = safe_float(get_text(cn,"nfe:vBC"))
                    valor = safe_float(get_text(cn,"nfe:vCOFINS"))
                    if aliq_str not in cof_por_aliq:
                        cof_por_aliq[aliq_str] = {"bc":0.0,"valor":0.0}
                    cof_por_aliq[aliq_str]["bc"]    += bc
                    cof_por_aliq[aliq_str]["valor"] += valor
                    break

    for aliq_str, dados in sorted(cof_por_aliq.items(), key=lambda x: safe_float(x[0])):
        if dados["valor"] > 0 or dados["bc"] > 0:
            linhas.append(r1020(5, base=fmt_decimal(str(dados["bc"])),
                                aliq=fmt_decimal(aliq_str,4),
                                valor=fmt_decimal(str(dados["valor"])), v_cont=v_nf))

    for aliq_str, dados in sorted(icms_por_aliq.items(), key=lambda x: safe_float(x[0])):
        if dados["deson"] > 0:
            linhas.append(r1020(45, base=fmt_decimal(str(dados["bc"])),
                                aliq=fmt_decimal(aliq_str),
                                valor=fmt_decimal(str(dados["deson"])), v_cont=v_nf))

    v_pis_tot    = get_text(total,"nfe:vPIS")
    v_cofins_tot = get_text(total,"nfe:vCOFINS")
    bc_pis_total = bc_cof_total = 0.0
    for det in det_list:
        imp = det.find("nfe:imposto", NS)
        if imp is None:
            continue
        for pt in ["PISAliq","PISQtde","PISNT","PISOutr"]:
            pn = imp.find(f"nfe:PIS/nfe:{pt}", NS)
            if pn is not None:
                try:
                    bc_pis_total += float(get_text(pn,"nfe:vBC") or "0")
                except ValueError:
                    pass
                break
        for ct in ["COFINSAliq","COFINSQtde","COFINSNT","COFINSOutr"]:
            cn = imp.find(f"nfe:COFINS/nfe:{ct}", NS)
            if cn is not None:
                try:
                    bc_cof_total += float(get_text(cn,"nfe:vBC") or "0")
                except ValueError:
                    pass
                break

    if v_pis_tot and safe_float(v_pis_tot) > 0:
        linhas.append(r1020(133, base=fmt_decimal(str(bc_pis_total)),
                            valor=fmt_decimal(v_pis_tot), v_cont=v_nf))
    if v_cofins_tot and safe_float(v_cofins_tot) > 0:
        linhas.append(r1020(134, base=fmt_decimal(str(bc_cof_total)),
                            valor=fmt_decimal(v_cofins_tot), v_cont=v_nf))
    return linhas

# ─────────────────────────────────────────────
# MUDANÇA 4/5 — gerar_registro_1030: normalizar CST + propagar tem_direito_credito
# ─────────────────────────────────────────────
def gerar_registro_1030(det, seq: int,
                         importacao: bool = False,
                         aliq_pis_pad: float = 0.0,
                         aliq_cof_pad: float = 0.0,
                         tem_direito_credito: bool = True) -> str:
    prod    = det.find("nfe:prod", NS)
    imposto = det.find("nfe:imposto", NS)
    cod_prod = get_text(prod,"nfe:cProd")[:14]
    qtd      = fmt_decimal(get_text(prod,"nfe:qCom"), 2)
    v_prod   = get_text(prod,"nfe:vProd")
    v_outro  = get_text(prod,"nfe:vOutro")
    v_desc   = get_text(prod,"nfe:vDesc")
    cfop     = get_text(prod,"nfe:CFOP")
    unidade  = get_text(prod,"nfe:uCom")
    v_unit   = get_text(prod,"nfe:vUnCom")
    cest     = get_text(prod,"nfe:CEST")
    di_node  = prod.find("nfe:DI", NS)
    n_di = somente_numeros(get_text(di_node,"nfe:nDI")) if di_node is not None else ""
    d_di = fmt_date(get_text(di_node,"nfe:dDI"))        if di_node is not None else ""
    icms_node = None
    v_bc_icms = aliq_icms = v_icms = cst_icms = v_icms_des = v_bc_st = ""
    v_ipi = aliq_ipi = cst_ipi = ""
    v_pis = aliq_pis = cst_pis_xml = bc_pis = ""
    v_cof = aliq_cof = cst_cof_xml = bc_cof = ""
    ibs_class_trib = ibs_bc = ibs_aliq = ibs_val = ""
    cbs_class_trib = cbs_bc = cbs_aliq = cbs_val = ""

    if imposto is not None:
        for tp in ["ICMS00","ICMS10","ICMS20","ICMS30","ICMS40","ICMS51","ICMS60",
                   "ICMS70","ICMS90","ICMSSN101","ICMSSN102","ICMSSN201",
                   "ICMSSN202","ICMSSN500","ICMSSN900"]:
            node = imposto.find(f"nfe:ICMS/nfe:{tp}", NS)
            if node is not None:
                icms_node = node
                break
        if icms_node is not None:
            v_bc_icms  = fmt_decimal(get_text(icms_node,"nfe:vBC"))
            aliq_icms  = fmt_decimal(get_text(icms_node,"nfe:pICMS"))
            v_icms     = fmt_decimal(get_text(icms_node,"nfe:vICMS"))
            cst_icms   = get_text(icms_node,"nfe:CST") or get_text(icms_node,"nfe:CSOSN")
            v_icms_des = fmt_decimal(get_text(icms_node,"nfe:vICMSDeson"))
            v_bc_st    = fmt_decimal(get_text(icms_node,"nfe:vBCST"))

        ipi_trib = imposto.find("nfe:IPI/nfe:IPITrib", NS)
        ipi_nt   = imposto.find("nfe:IPI/nfe:IPINT", NS)
        if ipi_trib is not None:
            v_ipi    = fmt_decimal(get_text(ipi_trib,"nfe:vIPI"))
            aliq_ipi = fmt_decimal(get_text(ipi_trib,"nfe:pIPI"))
            cst_ipi  = get_text(ipi_trib,"nfe:CST")
        elif ipi_nt is not None:
            v_ipi = "0,00"; aliq_ipi = "0,00"
            cst_ipi = get_text(ipi_nt,"nfe:CST")

        # ── MUDANÇA 4: normalizar CST PIS com zfill(2) antes de obter_cst_entrada ──
        pis_node = imposto.find("nfe:PIS", NS)
        if pis_node is not None:
            for pt in ["PISAliq","PISQtde","PISNT","PISOutr"]:
                pn = pis_node.find(f"nfe:{pt}", NS)
                if pn is not None:
                    # NORMALIZAÇÃO zfill(2)
                    cst_pis_xml = str(get_text(pn,"nfe:CST")).strip().zfill(2)
                    aliq_raw    = get_text(pn,"nfe:pPIS") or get_text(pn,"nfe:vAliqProd")
                    v_pis       = fmt_decimal(get_text(pn,"nfe:vPIS"))
                    aliq_pis    = fmt_decimal(aliq_raw, 4)
                    bc_pis      = fmt_decimal(get_text(pn,"nfe:vBC"))
                    break

        # ── MUDANÇA 4: normalizar CST COFINS com zfill(2) antes de obter_cst_entrada ──
        cof_node = imposto.find("nfe:COFINS", NS)
        if cof_node is not None:
            for ct in ["COFINSAliq","COFINSQtde","COFINSNT","COFINSOutr"]:
                cn = cof_node.find(f"nfe:{ct}", NS)
                if cn is not None:
                    # NORMALIZAÇÃO zfill(2)
                    cst_cof_xml = str(get_text(cn,"nfe:CST")).strip().zfill(2)
                    aliq_raw    = get_text(cn,"nfe:pCOFINS") or get_text(cn,"nfe:vAliqProd")
                    v_cof       = fmt_decimal(get_text(cn,"nfe:vCOFINS"))
                    aliq_cof    = fmt_decimal(aliq_raw, 4)
                    bc_cof      = fmt_decimal(get_text(cn,"nfe:vBC"))
                    break

        ibs_node = imposto.find("nfe:IBSCBS", NS)
        if ibs_node is not None:
            ibs_class_trib = get_text(ibs_node,"nfe:cClassTrib")
            cbs_class_trib = ibs_class_trib
            gibs = ibs_node.find("nfe:gIBSCBS", NS)
            if gibs is not None:
                ibs_bc = fmt_decimal(get_text(gibs,"nfe:vBC"))
                cbs_bc = ibs_bc
                guf = gibs.find("nfe:gIBSUF", NS)
                if guf is not None:
                    ibs_aliq = fmt_decimal(get_text(guf,"nfe:pIBSUF"))
                    ibs_val  = fmt_decimal(get_text(guf,"nfe:vIBSUF"))
                gcbs = gibs.find("nfe:gCBS", NS)
                if gcbs is not None:
                    cbs_aliq = fmt_decimal(get_text(gcbs,"nfe:pCBS"))
                    cbs_val  = fmt_decimal(get_text(gcbs,"nfe:vCBS"))

    # MUDANÇA 4: usar obter_cst_entrada com CST já normalizado
    cst_pis_ef = obter_cst_entrada(cst_pis_xml, tem_direito_credito)
    cst_cof_ef = obter_cst_entrada(cst_cof_xml, tem_direito_credito)

    try:
        vp = float(v_prod or "0")
        vi = float(
            get_text(
                imposto.find("nfe:IPI/nfe:IPITrib", NS) if imposto else None,
                "nfe:vIPI"
            ) or "0"
        ) if imposto else 0.0
        v_total = fmt_decimal(str(vp + vi))
    except (ValueError, TypeError):
        v_total = fmt_decimal(v_prod)

    vinculo_pis = "08" if importacao else ""
    vinculo_cof = "08" if importacao else ""

    c = [""] * 111
    c[0]="1030"; c[1]=cod_prod; c[2]=qtd; c[3]=v_total; c[4]=v_ipi
    c[5]=fmt_decimal(v_prod); c[6]="1"; c[7]=d_di; c[8]=n_di; c[9]=cst_icms
    c[10]=fmt_decimal(v_prod); c[11]=fmt_decimal(v_desc); c[12]=v_bc_icms
    c[13]=v_bc_st; c[14]=aliq_icms; c[15]=""; c[16]=""; c[17]=""; c[18]=""
    c[19]=fmt_decimal(v_outro); c[20]=""; c[21]=v_icms; c[22]=""; c[23]=""
    c[24]=""; c[25]=""; c[26]=fmt_decimal(v_unit,6); c[27]=""; c[28]=cst_ipi
    c[29]=aliq_ipi; c[30]=""; c[31]=""; c[32]=""; c[33]=cfop; c[34]=""
    c[35]=aliq_pis; c[36]=v_pis; c[37]=aliq_cof; c[38]=v_cof
    c[39]=fmt_decimal(v_prod); c[40]=cst_pis_ef; c[41]=bc_pis  # campo 41 = CST PIS efetivo
    c[42]=cst_cof_ef; c[43]=bc_cof                              # campo 43 = CST COF efetivo
    c[44]=""; c[45]=""; c[46]=""; c[47]=""
    c[48]=""; c[49]=""; c[50]=""; c[51]=""; c[52]=""; c[53]=""; c[54]=""
    c[55]="S"; c[56]=unidade; c[57]=""; c[58]=""; c[59]=fmt_decimal(v_prod)
    c[60]=""; c[61]=""; c[62]=""; c[63]=""; c[64]=""; c[65]=""; c[66]=""
    c[67]=""; c[68]=""; c[69]=""; c[70]=""
    c[71]=vinculo_pis; c[72]=vinculo_cof
    c[73]=""; c[74]=""; c[75]=""; c[76]=""; c[77]=""; c[78]=""; c[79]=""
    c[80]=""; c[81]=""; c[82]=""; c[83]=""; c[84]=""; c[85]=""; c[86]=""
    c[87]=""; c[88]=""; c[89]=""; c[90]=cest; c[91]=""; c[92]=""; c[93]=""
    c[94]=""; c[95]=""; c[96]=v_icms_des; c[97]=""; c[98]=""; c[99]=""
    c[100]=""; c[101]=""; c[102]=""
    c[103]=ibs_class_trib; c[104]=ibs_bc; c[105]=ibs_aliq; c[106]=ibs_val
    c[107]=cbs_class_trib; c[108]=cbs_bc; c[109]=cbs_aliq; c[110]=cbs_val
    return pipe_join(c)

def gerar_registro_1097(nfe_root) -> str:
    transp = nfe_root.find("nfe:infNFe/nfe:transp", NS)
    if transp is None:
        return ""
    mod_frete_cod = get_text(transp,"nfe:modFrete")
    mod_frete     = {"0":"C","1":"F","2":"S","3":"T","4":"R","5":"D","9":"S"}.get(mod_frete_cod,"C")
    frete_conta   = "E" if mod_frete == "C" else "D"
    det_list = nfe_root.findall("nfe:infNFe/nfe:det", NS)
    tp_via = ""
    if det_list:
        di_node = det_list[0].find("nfe:prod/nfe:DI", NS)
        if di_node is not None:
            tp_via = get_text(di_node,"nfe:tpViaTransp")
    transporta   = transp.find("nfe:transporta", NS)
    cnpj_transp  = get_text(transporta,"nfe:CNPJ")   if transporta is not None else ""
    razao_transp = get_text(transporta,"nfe:xNome")  if transporta is not None else ""
    ie_transp    = get_text(transporta,"nfe:IE")     if transporta is not None else ""
    end_transp   = get_text(transporta,"nfe:xEnder") if transporta is not None else ""
    uf_transp    = get_text(transporta,"nfe:UF")     if transporta is not None else ""
    cmun_transp  = get_text(transporta,"nfe:cMun")   if transporta is not None else ""
    cidade_cod   = somente_numeros(cmun_transp) if cmun_transp else ""
    tipo_insc    = "1" if cnpj_transp else ""
    vol     = transp.find("nfe:vol", NS)
    q_vol   = get_text(vol,"nfe:qVol")  if vol is not None else ""
    esp_vol = get_text(vol,"nfe:esp")   if vol is not None else ""
    marca   = get_text(vol,"nfe:marca") if vol is not None else ""
    peso_l  = fmt_decimal(get_text(vol,"nfe:pesoL"),3) if vol is not None else ""
    peso_b  = fmt_decimal(get_text(vol,"nfe:pesoB"),3) if vol is not None else ""
    return pipe_join([
        "1097", mod_frete, tp_via, frete_conta, "", "", "", "", "", "",
        razao_transp[:150] if razao_transp else "",
        tipo_insc, cnpj_transp, ie_transp, end_transp,
        "", "", "", cidade_cod, uf_transp, "",
        q_vol, esp_vol, marca, "", peso_l, peso_b,
        "E", "", "", "", "D", "", "", "",
    ])

def gerar_registro_1150(ct, bc, aliq, valor) -> str:
    return pipe_join(["1150", ct, bc, aliq, valor])

def gerar_registro_1151(ct, bc, aliq, valor) -> str:
    return pipe_join(["1151", ct, bc, aliq, valor])

# ─────────────────────────────────────────────
# CONVERSÃO PRINCIPAL
# MUDANÇA 5 — propagar tem_direito_credito em converter_xml e no loop de itens
# ─────────────────────────────────────────────
def converter_xml(xml_content: bytes, cnpj_fallback: str,
                  acumulador: str = "1157", especie: str = "36",
                  incluir_0000: bool = True, incluir_0020: bool = True,
                  incluir_0100: bool = True, incluir_0110: bool = True,
                  incluir_1010: bool = True, incluir_1015: bool = False,
                  incluir_1097: bool = True, grupo_padrao: int = 0,
                  tem_direito_credito: bool = True) -> tuple:

    root, erro = parse_xml_seguro(xml_content)
    if root is None:
        return "", {"erro": erro}, []

    nfe = root.find("nfe:NFe", NS)
    if nfe is None:
        nfe = root
    importacao = is_nota_importacao(nfe)
    cnpj_empresa, origem_cnpj = extrair_cnpj_empresa(nfe, cnpj_fallback)
    if not cnpj_empresa:
        return "", {"erro": "CNPJ nao encontrado. Para importacao, informe o CNPJ Fallback."}, []

    aliq_pis_pad, aliq_cof_pad = calcular_aliquotas_padrao_nota(nfe)
    lines    = []
    det_list = nfe.findall("nfe:infNFe/nfe:det", NS)
    emit     = nfe.find("nfe:infNFe/nfe:emit", NS)
    dest_node= nfe.find("nfe:infNFe/nfe:dest", NS)
    ide      = nfe.find("nfe:infNFe/nfe:ide", NS)
    total    = nfe.find("nfe:infNFe/nfe:total/nfe:ICMSTot", NS)

    if importacao and dest_node is not None:
        nome_forn = get_text(dest_node,"nfe:xNome"); uf_forn = "EX"
        ender_dest = dest_node.find("nfe:enderDest", NS)
        c_pais_xml = get_text(ender_dest,"nfe:cPais") if ender_dest is not None else ""
        x_pais_xml = get_text(ender_dest,"nfe:xPais") if ender_dest is not None else ""
        cod_pais_dominio = resolver_codigo_pais_dominio(c_pais_xml, x_pais_xml)
    else:
        nome_forn = get_text(emit,"nfe:xNome") if emit is not None else ""
        uf_forn = ""; cod_pais_dominio = ""
        if emit is not None:
            ender_e = emit.find("nfe:enderEmit", NS)
            uf_forn = get_text(ender_e,"nfe:UF") if ender_e is not None else ""

    chave_resumo = extrair_chave_nfe(nfe)
    resumo = {
        "nNF":           get_text(ide,"nfe:nNF"),
        "Emitente":      get_text(emit,"nfe:xNome") if emit is not None else "",
        "CNPJ Emit":     get_text(emit,"nfe:CNPJ")  if emit is not None else "",
        "Fornecedor":    nome_forn,
        "UF Forn":       uf_forn,
        "CNPJ Empresa":  cnpj_empresa,
        "Origem CNPJ":   origem_cnpj,
        "Importacao":    "Sim" if importacao else "Nao",
        "Emitente NF":   "P (Proprio)" if importacao else "T (Terceiros)",
        "Emissao":       fmt_date(get_text(ide,"nfe:dhEmi")),
        "Itens":         len(det_list),
        "vNF":           fmt_decimal(get_text(total,"nfe:vNF")),
        "vICMS":         fmt_decimal(get_text(total,"nfe:vICMS")),
        "vICMSDes":      fmt_decimal(get_text(total,"nfe:vICMSDeson")),
        "vIPI":          fmt_decimal(get_text(total,"nfe:vIPI")),
        "vPIS":          fmt_decimal(get_text(total,"nfe:vPIS")),
        "vCOFINS":       fmt_decimal(get_text(total,"nfe:vCOFINS")),
        "Chave NF-e":    chave_resumo,
        "Cod Pais (Dom)":cod_pais_dominio,
        "Aliq PIS Pad":  fmt_decimal(str(aliq_pis_pad),4),
        "Aliq COF Pad":  fmt_decimal(str(aliq_cof_pad),4),
        "Grupo":         (f"{grupo_padrao} - {TABELA_GRUPOS.get(grupo_padrao,'GERAL')}"
                          if grupo_padrao > 0 else "Auto (CFOP/NCM)"),
        "Direito Credito": "Sim" if tem_direito_credito else "Nao",
    }

    if incluir_0000:
        lines.append(gerar_registro_0000(cnpj_empresa))
    if incluir_0020 and emit is not None:
        lines.append(gerar_registro_0020(emit, dest=dest_node, is_importacao=importacao))
    if incluir_0100:
        produtos_gerados = set()
        for det in det_list:
            cod = get_text(det.find("nfe:prod",NS),"nfe:cProd")
            if cod not in produtos_gerados:
                lines.append(gerar_registro_0100(det, grupo_padrao=grupo_padrao))
                if incluir_0110:
                    lines.append(gerar_registro_0110(
                        det,
                        importacao=importacao,
                        aliq_pis_pad=aliq_pis_pad,
                        aliq_cof_pad=aliq_cof_pad,
                        tem_direito_credito=tem_direito_credito,
                    ))
                produtos_gerados.add(cod)

    lines.append(gerar_registro_1000(nfe, cnpj_empresa, acumulador, especie, importacao))

    if incluir_1010:
        for r in gerar_registros_1010(nfe):
            lines.append(r)
    if incluir_1015:
        for r in gerar_registros_1015(nfe):
            lines.append(r)

    for r in gerar_registros_1020(nfe, importacao):
        lines.append(r)

    for seq, det in enumerate(det_list, start=1):
        lines.append(gerar_registro_1030(
            det, seq,
            importacao=importacao,
            aliq_pis_pad=aliq_pis_pad,
            aliq_cof_pad=aliq_cof_pad,
            tem_direito_credito=tem_direito_credito,
        ))

    if incluir_1097:
        r1097 = gerar_registro_1097(nfe)
        if r1097:
            lines.append(r1097)

    ibs_gerados = {}
    for det in det_list:
        imp = det.find("nfe:imposto", NS)
        if imp is None:
            continue
        ibs_node = imp.find("nfe:IBSCBS", NS)
        if ibs_node is None:
            continue
        ct   = get_text(ibs_node,"nfe:cClassTrib")
        gibs = ibs_node.find("nfe:gIBSCBS", NS)
        if not ct or gibs is None:
            continue
        if ct not in ibs_gerados:
            ibs_gerados[ct] = {"bc_ibs":0.0,"v_ibs":0.0,"aliq_ibs":"",
                               "bc_cbs":0.0,"v_cbs":0.0,"aliq_cbs":""}
        try:
            ibs_gerados[ct]["bc_ibs"] += float(get_text(gibs,"nfe:vBC") or "0")
        except ValueError:
            pass
        guf = gibs.find("nfe:gIBSUF", NS)
        if guf is not None:
            try:
                ibs_gerados[ct]["v_ibs"]   += float(get_text(guf,"nfe:vIBSUF") or "0")
                ibs_gerados[ct]["aliq_ibs"] = get_text(guf,"nfe:pIBSUF")
            except ValueError:
                pass
        gcbs = gibs.find("nfe:gCBS", NS)
        if gcbs is not None:
            try:
                ibs_gerados[ct]["bc_cbs"]  += float(get_text(gibs,"nfe:vBC") or "0")
                ibs_gerados[ct]["v_cbs"]   += float(get_text(gcbs,"nfe:vCBS") or "0")
                ibs_gerados[ct]["aliq_cbs"] = get_text(gcbs,"nfe:pCBS")
            except ValueError:
                pass

    for ct, d in ibs_gerados.items():
        lines.append(gerar_registro_1150(
            ct, fmt_decimal(str(d["bc_ibs"])),
            fmt_decimal(d["aliq_ibs"]), fmt_decimal(str(d["v_ibs"]))
        ))
        lines.append(gerar_registro_1151(
            ct, fmt_decimal(str(d["bc_cbs"])),
            fmt_decimal(d["aliq_cbs"]), fmt_decimal(str(d["v_cbs"]))
        ))

    return "\n".join(lines), resumo, (aliq_pis_pad, aliq_cof_pad)

# ─────────────────────────────────────────────
# SIDEBAR
# MUDANÇA 2 — adicionar tem_direito_credito logo após o campo acumulador
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("### Info")
    st.markdown(f"**Versao:** {VERSAO}")
    st.markdown("**Thomson Reuters | Dominio Sistemas**")
    st.markdown("---")
    st.markdown("### Parametros")
    cnpj_fallback = st.text_input(
        "CNPJ da Empresa (obrigatorio para importacao)",
        value="", max_chars=14
    )
    acumulador = st.text_input("Codigo do Acumulador (fallback)", value="1157")

    # ── MUDANÇA 2: checkbox tem_direito_credito ──────────────────────────
    tem_direito_credito = st.checkbox(
        "Tem direito a crédito de PIS/COFINS",
        value=True,
        help=(
            "Marcado (padrão): CST de saída (01-09, 49, 99) é convertido para o "
            "CST de entrada com direito a crédito (ex: 01→50, 06→73, 07→71). "
            "Desmarcado: o CST do XML é gravado sem transformação."
        ),
    )
    # ────────────────────────────────────────────────────────────────────

    especie = st.text_input("Codigo da Especie", value="36")
    st.markdown("---")
    st.markdown("### Registros de cadastro")
    inc_0000 = st.checkbox("0000 - Identificacao da empresa", value=True)
    inc_0020 = st.checkbox("0020 - Fornecedor", value=True)
    inc_0100 = st.checkbox("0100 - Produtos", value=True)
    inc_0110 = st.checkbox("0110 - Vigencia PIS/COFINS", value=True)
    st.markdown("---")
    st.markdown("### Registros filhos do 1000")
    inc_1010 = st.checkbox("1010 - Inf. Complementares", value=True)
    inc_1015 = st.checkbox("1015 - Observacoes", value=False)
    inc_1097 = st.checkbox("1097 - Dados do Frete", value=True)
    st.caption("1020 / 1030 / 1150 / 1151 sempre gerados.")
    st.markdown("---")
    st.markdown("### Grupo de Produtos (0100 campo 9)")
    grupo_selecionado = st.selectbox(
        "Grupo",
        options=list(TABELA_GRUPOS.keys()),
        format_func=lambda x: f"{x} - {TABELA_GRUPOS[x]}" if x > 0 else TABELA_GRUPOS[x],
        index=0,
    )
    st.markdown("---")
    with st.expander("Tabela de Grupos"):
        for cod, desc in sorted(TABELA_GRUPOS.items()):
            if cod > 0:
                st.caption(f"`{cod:3d}` - {desc}")
    with st.expander("Tabela CST PIS/COFINS Entrada"):
        for cst, desc in sorted(TABELA_CST_PIS_COFINS_ENTRADA.items()):
            st.caption(f"`{cst}` - {desc}")

with st.expander("Instrucoes / Historico de versoes", expanded=False):
    st.markdown("""
        <div class="instrucoes-box">
        <h4>V5.1-FINAL — CST PIS/COFINS: tabelas completas + tem_direito_credito</h4>
        <ul>
          <li><b>TABELA_CST_PIS_COFINS_ENTRADA / SAIDA</b>: tabelas completas conforme legislacao.</li>
          <li><b>obter_cst_entrada()</b>: substitui cst_pis_efetivo/cst_cof_efetivo.
              Se tem_direito_credito=True converte CST de saida para entrada (01→50, 06→73 etc.).
              Se False, devolve o CST do XML sem transformacao.</li>
          <li><b>Sidebar</b>: checkbox "Tem direito a credito" logo apos o campo Acumulador.</li>
          <li><b>extrair_pis_cofins</b>: normaliza CST com zfill(2) antes de qualquer logica.</li>
          <li><b>gerar_registro_1030</b>: normaliza CST PIS/COFINS com zfill(2) e usa obter_cst_entrada.</li>
          <li><b>extrair_dados_impostos_itens</b>: mesma normalizacao + obter_cst_entrada para o Excel.</li>
          <li>Coluna "PIS c/ Credito" / "COF c/ Credito" no Excel indica conversao aplicada (verde).</li>
        </ul>
        <h4>V5.0-FINAL — Correcao definitiva do erro "no element found: line 1, column 0"</h4>
        <h4>V4.9-FINAL — Correcao dos campos do registro 1000</h4>
        <h4>V4.6-FINAL — CST 73 + Excel + maiusculas</h4>
        </div>
    """, unsafe_allow_html=True)

st.markdown("---")
st.markdown("#### Upload de arquivos")
st.caption("Aceita **XML** individuais ou **ZIP** (somente XMLs com CFOP de importacao serao processados).")

uploaded_files = st.file_uploader(
    "Selecione arquivos XML ou um arquivo ZIP",
    type=["xml","zip"],
    accept_multiple_files=True,
)

if uploaded_files:
    arquivos_para_processar = []
    relatorio_zip           = []
    for f in uploaded_files:
        nome_lower = f.name.lower()
        if nome_lower.endswith(".zip"):
            zip_bytes = f.read()
            xmls_imp, total_xml, ignorados, erros_parse = extrair_xmls_importacao_do_zip(zip_bytes)
            relatorio_zip.append({
                "zip": f.name, "total": total_xml,
                "importacao": len(xmls_imp), "ignorados": ignorados, "erros": erros_parse
            })
            for item in xmls_imp:
                arquivos_para_processar.append(item)
        elif nome_lower.endswith(".xml"):
            raw = f.read()
            if raw:
                arquivos_para_processar.append({
                    "nome": f.name,
                    "bytes": sanitizar_xml_bytes(raw)
                })
            else:
                relatorio_zip.append({
                    "zip": f.name, "total": 0, "importacao": 0,
                    "ignorados": [], "erros": [f"{f.name} (arquivo vazio)"]
                })

    if relatorio_zip:
        st.markdown("#### Relatorio de triagem dos ZIPs")
        for rz in relatorio_zip:
            classe = "zip-info" if rz["importacao"] > 0 else "zip-warn"
            st.markdown(
                f'<div class="{classe}"><b>{rz["zip"]}</b> — '
                f'{rz["total"]} XML(s) | <b>{rz["importacao"]} importacao(s)</b> | '
                f'{len(rz["ignorados"])} ignorado(s) | {len(rz["erros"])} erro(s)</div>',
                unsafe_allow_html=True)
            if rz["ignorados"]:
                with st.expander(f"XMLs ignorados ({len(rz['ignorados'])})"):
                    for n in rz["ignorados"]:
                        st.caption(f"  {n}")
            if rz["erros"]:
                with st.expander(f"Erros de parse ({len(rz['erros'])})"):
                    for n in rz["erros"]:
                        st.caption(f"  {n}")

    if not arquivos_para_processar:
        st.warning("Nenhum XML de importacao encontrado para processar.")
        st.stop()

    st.markdown("---")
    st.markdown("#### 🔁 DE/PARA: CFOP → Acumulador")

    cfops_detectadas = extrair_cfops_importacao_dos_arquivos(arquivos_para_processar)

    if "cfop_acumulador_map" not in st.session_state:
        st.session_state["cfop_acumulador_map"] = {}

    for cfop in cfops_detectadas:
        if cfop not in st.session_state["cfop_acumulador_map"]:
            st.session_state["cfop_acumulador_map"][cfop] = CFOP_ACUMULADOR_PADRAO.get(cfop, "")

    editar_acumuladores = st.toggle(
        "✏️ Alterar acumuladores padrao", value=False,
        help="Ative para editar o acumulador de cada CFOP de importacao detectada nos XMLs."
    )

    if cfops_detectadas:
        cfops_sem_acum = [
            c for c in cfops_detectadas
            if not st.session_state["cfop_acumulador_map"].get(c, "").strip()
        ]
        if editar_acumuladores:
            if cfops_sem_acum:
                st.warning(f"⚠️ {len(cfops_sem_acum)} CFOP(s) sem acumulador: "
                           f"**{', '.join(cfops_sem_acum)}** — preencha antes de processar.")
            else:
                st.success("✅ Todas as CFOPs detectadas possuem acumulador definido.")

            h1, h2, h3, h4 = st.columns([1, 4, 2, 2])
            h1.markdown("**CFOP**"); h2.markdown("**Descricao**")
            h3.markdown("**Acumulador**"); h4.markdown("**Status**")
            st.markdown("<hr style='margin:4px 0 8px 0; border-color:#FF8000;'>",
                        unsafe_allow_html=True)

            novo_mapa = dict(st.session_state["cfop_acumulador_map"])
            for cfop in cfops_detectadas:
                desc   = CFOP_DESCRICAO.get(cfop, "CFOP nao catalogada")
                padrao = CFOP_ACUMULADOR_PADRAO.get(cfop, None)
                atual  = st.session_state["cfop_acumulador_map"].get(cfop, "")
                if not atual.strip():
                    status_txt, status_cor = "⚠️ Nao definido", "#C62828"
                elif padrao is not None and atual == padrao and padrao != "":
                    status_txt, status_cor = "✅ Padrao", "#2E7D32"
                else:
                    status_txt, status_cor = "✏️ Alterado", "#E65100"
                c1, c2, c3, c4 = st.columns([1, 4, 2, 2])
                c1.markdown(
                    f"<span style='font-family:monospace;font-weight:bold;color:#FF8000;'>{cfop}</span>",
                    unsafe_allow_html=True)
                c2.markdown(f"<small style='color:#555;'>{desc}</small>", unsafe_allow_html=True)
                novo_val = c3.text_input(
                    label=f"acum_{cfop}", value=atual, placeholder="Ex: 1157",
                    label_visibility="collapsed", key=f"acum_input_{cfop}", max_chars=10)
                novo_mapa[cfop] = novo_val.strip()
                padrao_label = padrao if (padrao is not None and padrao != "") else "sem padrao"
                c4.markdown(
                    f"<small style='color:{status_cor};font-weight:bold;'>{status_txt}</small>"
                    f"<br><small style='color:#888;'>padrao: {padrao_label}</small>",
                    unsafe_allow_html=True)
            st.session_state["cfop_acumulador_map"] = novo_mapa
            st.markdown("")
            col_reset, _ = st.columns([2, 6])
            with col_reset:
                if st.button("↩️ Restaurar todos os padroes", use_container_width=True):
                    for cfop in cfops_detectadas:
                        st.session_state["cfop_acumulador_map"][cfop] = \
                            CFOP_ACUMULADOR_PADRAO.get(cfop, "")
                    st.rerun()
        else:
            if cfops_sem_acum:
                st.warning(f"⚠️ {len(cfops_sem_acum)} CFOP(s) sem acumulador: "
                           f"**{', '.join(cfops_sem_acum)}** — ative '✏️ Alterar' para preencher.")
            h1, h2, h3, h4 = st.columns([1, 4, 2, 2])
            h1.markdown("**CFOP**"); h2.markdown("**Descricao**")
            h3.markdown("**Acumulador**"); h4.markdown("**Status**")
            st.markdown("<hr style='margin:4px 0 8px 0; border-color:#FF8000;'>",
                        unsafe_allow_html=True)
            for cfop in cfops_detectadas:
                desc       = CFOP_DESCRICAO.get(cfop, "CFOP nao catalogada")
                padrao     = CFOP_ACUMULADOR_PADRAO.get(cfop, None)
                acum_atual = st.session_state["cfop_acumulador_map"].get(cfop, "")
                if not acum_atual.strip():
                    status_txt, status_cor = "⚠️ Nao definido", "#C62828"
                    acum_display = "<span style='color:#C62828;'>—</span>"
                elif padrao is not None and acum_atual == padrao and padrao != "":
                    status_txt, status_cor = "✅ Padrao", "#2E7D32"
                    acum_display = f"<b style='font-size:15px;'>{acum_atual}</b>"
                else:
                    status_txt, status_cor = "✏️ Alterado", "#E65100"
                    acum_display = f"<b style='font-size:15px;'>{acum_atual}</b>"
                c1, c2, c3, c4 = st.columns([1, 4, 2, 2])
                c1.markdown(
                    f"<span style='font-family:monospace;font-weight:bold;color:#FF8000;'>{cfop}</span>",
                    unsafe_allow_html=True)
                c2.markdown(f"<small style='color:#555;'>{desc}</small>", unsafe_allow_html=True)
                c3.markdown(acum_display, unsafe_allow_html=True)
                c4.markdown(
                    f"<small style='color:{status_cor};font-weight:bold;'>{status_txt}</small>",
                    unsafe_allow_html=True)
    else:
        st.caption("Nenhuma CFOP de importacao (3xxx) detectada nos arquivos carregados.")

    st.markdown("---")

    mapa_acum = st.session_state.get("cfop_acumulador_map", {})
    cfops_sem_acum_final = [
        c for c in cfops_detectadas
        if not mapa_acum.get(c, "").strip()
    ]
    if cfops_sem_acum_final:
        st.error(
            f"❌ Nao e possivel processar: CFOPs **{', '.join(cfops_sem_acum_final)}** "
            f"sem acumulador. Ative '✏️ Alterar acumuladores padrao' e preencha."
        )
        st.stop()

    # ── PROCESSAMENTO ────────────────────────────────────────────────────
    all_lines = []; all_resumos = []; all_dados_xls = []; erros = []
    progress = st.progress(0, text="Processando arquivos...")

    for i, arq in enumerate(arquivos_para_processar):
        root_tmp, _ = parse_xml_seguro(arq["bytes"])
        if root_tmp is not None:
            nfe_tmp = root_tmp.find("nfe:NFe", NS) or root_tmp
            try:
                aliq_pis_p, aliq_cof_p = calcular_aliquotas_padrao_nota(nfe_tmp)
                # MUDANÇA 5: propagar tem_direito_credito para o Excel
                all_dados_xls.extend(
                    extrair_dados_impostos_itens(
                        nfe_tmp, arq["nome"],
                        aliq_pis_p, aliq_cof_p,
                        tem_direito_credito=tem_direito_credito,
                    )
                )
            except Exception:
                pass

            acumulador_nota = acumulador
            try:
                det_list_tmp = nfe_tmp.findall("nfe:infNFe/nfe:det", NS)
                if det_list_tmp:
                    cfop_primeiro = get_text(
                        det_list_tmp[0].find("nfe:prod", NS), "nfe:CFOP"
                    ).strip()
                    acumulador_nota = resolver_acumulador_por_cfop(
                        cfop_primeiro, mapa_acum, acumulador
                    )
            except Exception:
                pass
        else:
            acumulador_nota = acumulador

        # MUDANÇA 5: propagar tem_direito_credito para converter_xml
        texto, resumo, _ = converter_xml(
            arq["bytes"], cnpj_fallback=cnpj_fallback,
            acumulador=acumulador_nota, especie=especie,
            incluir_0000=inc_0000, incluir_0020=inc_0020,
            incluir_0100=inc_0100, incluir_0110=inc_0110,
            incluir_1010=inc_1010, incluir_1015=inc_1015,
            incluir_1097=inc_1097, grupo_padrao=grupo_selecionado,
            tem_direito_credito=tem_direito_credito,
        )

        if "erro" in resumo:
            erros.append({"Arquivo": arq["nome"], "Erro": resumo["erro"]})
        else:
            resumo["Acumulador"] = acumulador_nota
            all_lines.append(texto)
            all_resumos.append({"Arquivo": arq["nome"], **resumo})

        progress.progress(
            (i + 1) / len(arquivos_para_processar),
            text=f"Processando {arq['nome']}..."
        )

    progress.empty()

    if erros:
        st.error("Erros encontrados:")
        st.dataframe(erros, use_container_width=True)

    if all_resumos:
        st.success(f"{len(all_resumos)} arquivo(s) convertido(s) com sucesso!")
        cnpjs_unicos = list({r["CNPJ Empresa"]:r for r in all_resumos}.values())
        if cnpjs_unicos:
            st.markdown("#### Empresa / Fornecedor")
            cols = st.columns(min(len(cnpjs_unicos), 4))
            for idx, r in enumerate(cnpjs_unicos[:4]):
                is_imp = r.get("Importacao","Nao") == "Sim"
                cor    = "#1565C0" if is_imp else "#FF8000"
                pais_info = f" | Pais Dom.: {r.get('Cod Pais (Dom)','')}" if is_imp else ""
                cred_info = "✅ Com credito" if r.get("Direito Credito") == "Sim" else "❌ Sem credito"
                with cols[idx]:
                    st.markdown(
                        f'<div class="cnpj-badge" style="color:{cor};border-color:{cor};">'
                        f'Empresa: {r["CNPJ Empresa"]}</div>'
                        f'<div class="info-origem" style="border-left-color:{cor};">'
                        f'{r.get("Origem CNPJ","")}<br>'
                        f'{"Importacao | Forn: "+r.get("Fornecedor","")[:40]+pais_info if is_imp else "Forn: "+r.get("Fornecedor","")[:50]}'
                        f'<br><small>Aliq. PIS pad: {r.get("Aliq PIS Pad","")} | '
                        f'COFINS pad: {r.get("Aliq COF Pad","")}</small>'
                        f'<br><small>Acumulador: <b>{r.get("Acumulador","")}</b> | {cred_info}</small>'
                        f'<br><small>Chave: {r.get("Chave NF-e","")[:22]}...</small>'
                        f'</div>', unsafe_allow_html=True)

        st.markdown("---")
        with st.expander("Resumo das notas processadas", expanded=True):
            st.dataframe(all_resumos, use_container_width=True)

        saida_final = "\n".join(all_lines)
        with st.expander("Preview do arquivo gerado (primeiras 80 linhas)"):
            st.code("\n".join(saida_final.split("\n")[:80]), language="text")

        st.markdown("---")
        saida_ansi = encode_ansi(saida_final)
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="Baixar Arquivo Dominio (.TXT ANSI)",
                data=saida_ansi, file_name="importacao_dominio.txt",
                mime="text/plain", use_container_width=True, type="primary")
        with col2:
            if EXCEL_DISPONIVEL and all_dados_xls:
                excel_bytes = gerar_excel_relatorio(all_dados_xls)
                st.download_button(
                    label="Baixar Relatorio Excel (.XLSX)",
                    data=excel_bytes, file_name="relatorio_impostos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            elif not EXCEL_DISPONIVEL:
                st.warning("openpyxl nao instalado. Execute: pip install openpyxl")

        st.markdown("---")
        st.markdown("#### Estatisticas")
        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("Notas",  len(all_resumos))
        c2.metric("Itens",  sum(r.get("Itens",0) for r in all_resumos))
        total_linhas = len([l for l in saida_final.split("\n") if l.startswith("|")])
        c3.metric("Linhas geradas", total_linhas)
        c4.metric("Erros", len(erros))
        try:
            total_nf = sum(
                float(r.get("vNF","0").replace(",","."))
                for r in all_resumos if r.get("vNF")
            )
            c5.metric(
                "Total NF (R$)",
                f"{total_nf:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            )
        except Exception:
            c5.metric("Total NF", "-")
else:
    st.info("Faca o upload de um ou mais arquivos XML ou de um arquivo ZIP contendo XMLs de NF-e.")

st.markdown("---")
st.caption(
    f"Conversor XML NF-e → Dominio Sistemas | Thomson Reuters | Python + Streamlit | {VERSAO}"
)
